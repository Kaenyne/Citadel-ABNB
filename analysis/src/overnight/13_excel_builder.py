"""Workstream 13, part 2: build `model/ABNB_driver_model.xlsx` with live Excel formulas.

Called by `13_driver_model.py`. Every forecast cell in the workbook is a formula off the `Inputs`
sheet or the `History` sheet; nothing downstream of `Inputs` is a pasted value. The three scenarios
are all computed in full (three blocks on Revenue / Costs / Cash), and the scenario selector on
`Inputs!$B$4` (named range `Scenario`) drives the `Active` column on `Inputs`, on `Valuation`
sections 1, 2, 3 and 4, and on `Card_5Nov`.

The `Recon` sheet puts the Python mirror's value next to each workbook formula so Excel itself
computes the difference when the file is opened.
"""
from __future__ import annotations

import importlib.util
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))

_spec = importlib.util.spec_from_file_location("dm13", os.path.join(HERE, "13_driver_model.py"))
DM = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(DM)

SCEN_COL = {"Bear": "C", "Base": "D", "Bull": "E"}          # Inputs sheet scenario columns
REGIONS = DM.REGIONS
FQ = DM.FQ

TITLE = Font(bold=True, size=14)
H1 = Font(bold=True, size=11, color="FFFFFF")
H2 = Font(bold=True, size=10)
SMALL = Font(size=8, italic=True, color="666666")
FILL_H = PatternFill("solid", fgColor="1F3864")
FILL_SEC = PatternFill("solid", fgColor="D9E1F2")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")           # yellow = hard input
FILL_OUT = PatternFill("solid", fgColor="E2EFDA")          # green = key output
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))

F_NUM = "#,##0"
F_NUM1 = "#,##0.0"
F_NUM2 = "#,##0.00"
F_PCT1 = "0.0%"
F_PCT2 = "0.00%"
F_X = '0.0"x"'


def col(n):
    return get_column_letter(n)


def put(ws, r, c, v, font=None, fmt=None, fill=None, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def section(ws, r, text, width=12):
    for c in range(1, width + 1):
        put(ws, r, c, None, fill=FILL_SEC)
    put(ws, r, 1, text, font=H2, fill=FILL_SEC)
    return r + 1


def header(ws, r, labels, start=1):
    for i, t in enumerate(labels):
        put(ws, r, start + i, t, font=H1, fill=FILL_H)
    return r + 1


# =====================================================================================================
def build_workbook(quarterly, annual, valrows, revdcf, grid, prices, ev_today):
    wb = Workbook()
    A = {(a["scenario"], a["year"]): a for a in annual}
    Q = {(q["scenario"], q["quarter"]): q for q in quarterly}

    # ------------------------------------------------------------------ INPUTS
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 10
    for c in "CDEF":
        ws.column_dimensions[c].width = 13
    ws.column_dimensions["G"].width = 110
    r = 1
    put(ws, r, 1, "ABNB driver model - Inputs", font=TITLE); r += 1
    put(ws, r, 1, "Overnight run 6-7 Sep 2026 (workstream 13). Yellow cells are inputs; everything "
                  "downstream is a live formula. The other hard numbers in the file are the History "
                  "actuals table, the external Street / Card bars (each with its vendor and date) and "
                  "the Recon sheet's Python-mirror column. Sources name the workstream that supplied "
                  "the value.", font=SMALL); r += 2

    put(ws, r, 1, "SCENARIO SELECTOR  (1 = Bear, 2 = Base, 3 = Bull)", font=H2)
    put(ws, r, 2, 2, fill=FILL_IN, font=Font(bold=True))
    SCEN_SEL = "Inputs!$B$%d" % r
    put(ws, r, 7, "Drives the Active column here and the Active column on Valuation (sections 1-4) "
                  "and Card_5Nov. All three scenarios are computed in full on Revenue / Costs / Cash, "
                  "so the Bear / Base / Bull columns never move when this cell changes.", font=SMALL)
    r += 2

    r = section(ws, r, "A. VALUATION AND CONTROL INPUTS")
    r = header(ws, r, ["Parameter", "Unit", "Bear", "Base", "Bull", "Active", "Source"])
    rv = {}
    for k in DM.VAL_ORDER:
        bear, base, bull, unit, src = DM.VAL_INPUTS[k]
        put(ws, r, 1, DM.VAL_LABEL[k])
        put(ws, r, 2, unit)
        fmt = F_PCT2 if k in ("cost_of_equity", "terminal_growth", "withholding_pct", "price_growth",
                              "nb_incr_margin", "dcf_start_growth", "margin_cap")             else (F_X if k.startswith("exit_") else F_NUM2)
        for i, v in enumerate((bear, base, bull)):
            put(ws, r, 3 + i, v, fmt=fmt, fill=FILL_IN)
        put(ws, r, 6, f"=CHOOSE({SCEN_SEL},C{r},D{r},E{r})", fmt=fmt)
        if k == "dcf_years":
            # WS17 audit: the Valuation DCF grid is written out as exactly ten rows, so this cell
            # documents the convention rather than resizing anything. Say so where a reader will see it.
            src += (". LAYOUT-FIXED: the DCF grid on Valuation section 3 is ten written-out rows, so "
                    "changing this cell does not resize the grid and has no effect on any output")
        put(ws, r, 7, src, font=SMALL)
        rv[k] = r
        r += 1
    r += 1

    r = section(ws, r, "B. ANNUAL DRIVERS BY FISCAL YEAR")
    ra = {}
    for yr in DM.YEARS:
        r = header(ws, r, [f"FY{yr}E driver", "Unit", "Bear", "Base", "Bull", "Active", "Source"])
        for k in DM.ANNUAL_ORDER:
            bear, base, bull, unit, src = DM.ANNUAL_INPUTS[yr][k]
            put(ws, r, 1, DM.ANNUAL_LABEL[k])
            put(ws, r, 2, unit)
            fmt = F_PCT2 if unit in ("y/y", "% rev", "% pretax") else (F_NUM1 if unit == "$M" else F_NUM2)
            for i, v in enumerate((bear, base, bull)):
                put(ws, r, 3 + i, v, fmt=fmt, fill=FILL_IN)
            put(ws, r, 6, f"=CHOOSE({SCEN_SEL},C{r},D{r},E{r})", fmt=fmt)
            put(ws, r, 7, src, font=SMALL)
            ra[(yr, k)] = r
            r += 1
        if yr == 2028:
            for k, tup, unit, src in [
                ("nights_2028", DM.NIGHTS_2028, "y/y",
                 "WS07 lever: no regional build exists three years out"),
                ("adr_exfx_2028", DM.ADR_EXFX_2028, "y/y", "WS07 lever"),
                # WS17 audit fix: these two are stored as fractions and percent-formatted, unlike the
                # quarterly FX rows in section C, which are plain percentage-point numbers divided by
                # 100 where they are used. The unit label used to read "pp" on both, so a reader who
                # typed 1.5 here meaning +1.5pp would have got +150%. Both are zero today.
                ("rev_fx_2028", tuple(DM.FX_2028[s][0] / 100 for s in DM.SCENARIOS), "y/y",
                 "WS07: no FX view three years out; zero in every case so FX does not flatter the "
                 "path. UNIT: a fraction, percent-formatted (0.015 = +1.5pp), NOT the plain "
                 "percentage-point number used by the quarterly FX rows in section C"),
                ("adr_fx_2028", tuple(DM.FX_2028[s][1] / 100 for s in DM.SCENARIOS), "y/y",
                 "same; a fraction, percent-formatted, not a percentage-point number"),
            ]:
                put(ws, r, 1, {"nights_2028": "Total nights growth, FY2028E",
                               "adr_exfx_2028": "ADR ex-FX growth, FY2028E",
                               "rev_fx_2028": "FX effect on revenue, FY2028E",
                               "adr_fx_2028": "FX effect on ADR, FY2028E"}[k])
                put(ws, r, 2, unit)
                for i, v in enumerate(tup):
                    put(ws, r, 3 + i, v, fmt=F_PCT2, fill=FILL_IN)
                put(ws, r, 6, f"=CHOOSE({SCEN_SEL},C{r},D{r},E{r})", fmt=F_PCT2)
                put(ws, r, 7, src, font=SMALL)
                ra[(2028, k)] = r
                r += 1
        r += 1

    r = section(ws, r, "C. QUARTERLY DRIVERS 3Q26 - 4Q27")
    r = header(ws, r, ["Quarter / driver", "Unit", "Bear", "Base", "Bull", "Active", "Source"])
    rq = {}
    QFIELDS = [("na", "North America nights growth", "y/y %"),
               ("emea", "EMEA nights growth", "y/y %"),
               ("latam", "Latin America nights growth", "y/y %"),
               ("apac", "Asia Pacific nights growth", "y/y %"),
               ("adr_exfx", "ADR ex-FX growth", "y/y %"),
               ("rev_fx", "FX effect on revenue", "pp"),
               ("adr_fx", "FX effect on ADR", "pp"),
               ("margin", "Adj. EBITDA margin (2026 quarters only)", "%")]
    for q in FQ:
        for key, label, unit in QFIELDS:
            if key == "margin" and q not in DM.Q_MARGIN_26:
                rq[(q, key)] = r          # row reserved but not used: 2027 margins are derived
                r += 1
                continue
            put(ws, r, 1, f"{q}  {label}")
            put(ws, r, 2, unit)
            for i, s in enumerate(DM.SCENARIOS):
                if key in REGIONS:
                    v = DM.REGIONAL_G[q][s][REGIONS.index(key)]
                    src = "WS10 10_regional_forecast.csv"
                elif key == "adr_exfx":
                    v = DM.ADR_EXFX_Q[q][s]
                    src = "WS10 (2026 quarters); WS06 and WS07 both +2.5% base for FY27"
                elif key == "rev_fx":
                    v = DM.FX_Q[q][s][0]
                    src = ("2Q26 letter guide (+3.0pp); WS05 lagged fit says +2.19pp" if q == "3Q26"
                           else "WS05 05_fx_schedule.csv, lagged fit; bear = strong_usd, bull = weak_usd")
                elif key == "adr_fx":
                    v = DM.FX_Q[q][s][1]
                    src = "WS05 05_fx_schedule.csv broad-USD ADR fit, re-estimated by WS08 (r 0.96)"
                else:
                    if q not in DM.Q_MARGIN_26:
                        continue
                    v = DM.Q_MARGIN_26[q][s]
                    src = ("WS07 5 Nov card: guide is 'down slightly' from 3Q25's 50.1%" if q == "3Q26"
                           else "WS07 implied-Q4 table")
                put(ws, r, 3 + i, v, fmt=F_NUM2, fill=FILL_IN)
            if not (key == "margin" and q not in DM.Q_MARGIN_26):
                put(ws, r, 6, f"=CHOOSE({SCEN_SEL},C{r},D{r},E{r})", fmt=F_NUM2)
                put(ws, r, 7, src, font=SMALL)
            rq[(q, key)] = r
            r += 1
    r += 1

    r = section(ws, r, "D. REGULATORY NIGHTS DRAG (WS11 median, incremental pp of y/y growth)")
    r = header(ws, r, ["Year / region", "Unit", "Value", "", "", "", "Source"])
    rd = {}
    for yr in DM.YEARS:
        for rg in REGIONS:
            put(ws, r, 1, f"FY{yr}  {DM.RLABEL[rg]}")
            put(ws, r, 2, "pp")
            put(ws, r, 3, DM.REG_DRAG_PP[yr][rg], fmt=F_NUM2, fill=FILL_IN)
            put(ws, r, 7, "WS11 11_regulatory_overlay.csv; cumulative run-rate loss vs the 2Q26 baseline, "
                          "differenced into a y/y growth drag. Scaled by the 'Regulatory drag multiplier' above.",
                font=SMALL)
            rd[(yr, rg)] = r
            r += 1
    r += 1

    r = section(ws, r, "E. DOCUMENTED ALTERNATIVES - values a workstream supplied that lost the base-case "
                       "argument. Not referenced by any formula; swap them in deliberately.")
    r = header(ws, r, ["Item", "Alternative (not used)", "In the base case", "", "", "", "Why the base wins"])
    for item, alt, base, why in DM.ALTERNATIVES:
        put(ws, r, 1, item, wrap=True)
        put(ws, r, 2, alt, wrap=True)
        put(ws, r, 3, base, wrap=True)
        put(ws, r, 7, why, font=SMALL, wrap=True)
        r += 1

    IN = lambda key_row, s: f"Inputs!${SCEN_COL[s]}${key_row}"
    VAL = lambda k, s: IN(rv[k], s)
    ANN = lambda yr, k, s: IN(ra[(yr, k)], s)
    QIN = lambda q, k, s: IN(rq[(q, k)], s)
    RDG = lambda yr, rg: f"Inputs!$C${rd[(yr, rg)]}"

    # ------------------------------------------------------------------ HISTORY
    hs = wb.create_sheet("History")
    hs.column_dimensions["A"].width = 40
    hcols = ["quarter", "nights_m", "gbv_musd", "adr_usd", "revenue_musd", "take_rate_pct",
             "adj_ebitda_musd", "adj_ebitda_margin_pct", "sbc_musd", "fcf_musd", "buybacks_musd",
             "rsu_tax_withholding_musd", "diluted_wa_shares_m", "net_income_musd",
             "cost_of_revenue_musd", "ops_support_musd", "product_dev_musd", "sales_marketing_musd",
             "g_and_a_musd", "exsbc__cor_cash", "exsbc__ops_cash", "exsbc__pd_cash", "exsbc__sm_cash",
             "exsbc__ga_cash"]
    hlab = ["Quarter", "Nights & experiences booked (M)", "GBV ($M)", "ADR ($)", "Revenue ($M)",
            "Take rate (%)", "Adj. EBITDA ($M)", "Adj. EBITDA margin (%)", "SBC ($M)", "FCF ($M)",
            "Buybacks ($M)", "RSU tax withholding ($M)", "Diluted WA shares (M)", "Net income ($M)",
            "Cost of revenue, GAAP ($M)", "Ops & support, GAAP ($M)", "Product development, GAAP ($M)",
            "Sales & marketing, GAAP ($M)", "G&A, GAAP ($M)", "Cost of revenue, cash ($M)",
            "Ops & support, cash ($M)", "Product development, cash ($M)", "Sales & marketing, cash ($M)",
            "G&A, cash ($M)"]
    r = 1
    put(hs, r, 1, "History - quarterly KPIs and cost lines, 1Q21 to 2Q26", font=TITLE); r += 1
    put(hs, r, 1, "Source: data/processed/overnight/02_kpi_panel_quarterly.csv (WS02), "
                  "data/processed/abnb_capital_return_quarterly.csv. Actuals, not formulas.", font=SMALL); r += 2
    hdr = r
    for i, t in enumerate(hlab):
        put(hs, r, 1 + i, t, font=H1, fill=FILL_H)
    r += 1
    qrow = {}
    for q in DM.HIST_Q:
        p = DM.PANEL[q]
        cr = DM.CAPRET.get(q, {})
        vals = []
        for c in hcols:
            if c == "quarter":
                vals.append(q)
            elif c in ("buybacks_musd", "rsu_tax_withholding_musd"):
                vals.append(DM.fl(cr.get(c)) if cr.get(c) else DM.fl(p.get(c)))
            else:
                vals.append(DM.fl(p.get(c)))
        for i, v in enumerate(vals):
            put(hs, r, 1 + i, v, fmt=(None if i == 0 else (F_NUM2 if hcols[i] in ("adr_usd", "take_rate_pct", "adj_ebitda_margin_pct") else F_NUM1)))
        qrow[q] = r
        r += 1
    for i in range(len(hlab)):
        hs.column_dimensions[col(1 + i)].width = 15 if i else 10

    r += 1
    r = section(hs, r, "MODEL ANCHORS - the actuals the forecast grows off (formulas over the table above)", 8)
    put(hs, r, 1, "Item", font=H1, fill=FILL_H); put(hs, r, 2, "Value", font=H1, fill=FILL_H)
    put(hs, r, 3, "Source / formula", font=H1, fill=FILL_H); r += 1
    anch = {}

    def anchor(label, formula, fmt=F_NUM1, note=""):
        nonlocal r
        put(hs, r, 1, label)
        put(hs, r, 2, formula, fmt=fmt)
        put(hs, r, 3, note, font=SMALL)
        anch[label] = f"History!$B${r}"
        r += 1

    fy25q = ["1Q25", "2Q25", "3Q25", "4Q25"]
    ci = {c: 1 + hcols.index(c) for c in hcols}
    S4 = lambda c, qs: "=" + "+".join(f"{col(ci[c])}{qrow[q]}" for q in qs)
    anchor("FY2025 nights (M)", S4("nights_m", fy25q))
    anchor("FY2025 GBV ($M)", S4("gbv_musd", fy25q))
    anchor("FY2025 revenue ($M)", S4("revenue_musd", fy25q))
    anchor("FY2025 ADR ($)", f"={anch['FY2025 GBV ($M)']}/{anch['FY2025 nights (M)']}", F_NUM2)
    anchor("FY2025 take rate", f"={anch['FY2025 revenue ($M)']}/{anch['FY2025 GBV ($M)']}", F_PCT2)
    anchor("FY2025 adj. EBITDA ($M)", S4("adj_ebitda_musd", fy25q))
    anchor("FY2025 SBC ($M)", S4("sbc_musd", fy25q))
    for lab, val, note in [
        ("FY2025 cost of revenue, cash ($M)", DM.FY25["cor"], "WS07 actuals()"),
        ("FY2025 ops & support, cash ($M)", DM.FY25["ops"], "WS07 actuals()"),
        ("FY2025 product development, cash ($M)", DM.FY25["pd"], "WS07 actuals()"),
        ("FY2025 brand & performance marketing ($M)", DM.FY25["bpm"], "WS07 07_cost_lines_per_night.csv"),
        ("FY2025 field operations & policy, cash ($M)", DM.FY25["fop"], "S&M cash less brand & performance"),
        ("FY2025 G&A, cash ($M)", DM.FY25["ga"], "WS07 actuals()"),
        ("FY2025 new-business revenue outside GBV x take ($M)", DM.FY25["new_business"],
         "WS11: Services $12M; sponsored listings did not exist"),
    ]:
        put(hs, r, 1, lab); put(hs, r, 2, val, fmt=F_NUM1, fill=FILL_IN); put(hs, r, 3, note, font=SMALL)
        anch[lab] = f"History!$B${r}"; r += 1
    h1q = ["1Q26", "2Q26"]
    anchor("1H2026 nights (M)", S4("nights_m", h1q), note="actual")
    anchor("1H2026 GBV ($M)", S4("gbv_musd", h1q), note="actual")
    anchor("1H2026 revenue ($M)", S4("revenue_musd", h1q), note="actual")
    anchor("1H2026 FCF ($M)", S4("fcf_musd", h1q), note="actual")
    anchor("1H2026 buybacks ($M)", S4("buybacks_musd", h1q), note="actual")
    anchor("1H2026 RSU withholding ($M)", S4("rsu_tax_withholding_musd", h1q), note="actual")
    anchor("1H2026 SBC ($M)", S4("sbc_musd", h1q), note="actual; used by the 2H26 share roll")
    for q in ["3Q25", "4Q25", "1Q26", "2Q26"]:
        anchor(f"{q} nights (M)", f"={col(ci['nights_m'])}{qrow[q]}")
        anchor(f"{q} GBV ($M)", f"={col(ci['gbv_musd'])}{qrow[q]}")
        anchor(f"{q} ADR ($)", f"={col(ci['adr_usd'])}{qrow[q]}", F_NUM2)
        anchor(f"{q} revenue ($M)", f"={col(ci['revenue_musd'])}{qrow[q]}")
        anchor(f"{q} take rate", f"={col(ci['take_rate_pct'])}{qrow[q]}/100", F_PCT2)
        for rg in REGIONS:
            put(hs, r, 1, f"{q} {DM.RLABEL[rg]} share of nights")
            put(hs, r, 2, DM.QSHARE[q][rg], fmt=F_PCT2, fill=FILL_IN)
            put(hs, r, 3, "WS10 10_regional_panel_quarterly.csv, normalised to 100%", font=SMALL)
            anch[f"{q} {rg} share"] = f"History!$B${r}"; r += 1
    for rg in REGIONS:
        put(hs, r, 1, f"FY2025 {DM.RLABEL[rg]} share of nights")
        put(hs, r, 2, DM.FY25_SHARE[rg], fmt=F_PCT2, fill=FILL_IN)
        put(hs, r, 3, "nights-weighted mean of the four 2025 quarters (WS10)", font=SMALL)
        anch[f"FY2025 {rg} share"] = f"History!$B${r}"; r += 1
    for qn in ["1Q", "2Q", "3Q", "4Q"]:
        put(hs, r, 1, f"Memo: seasonal adj. EBITDA margin spread, {qn}, 2023-2025 mean (pts vs the full year)")
        put(hs, r, 2, DM.SEAS_MARGIN[qn] / 100.0, fmt=F_PCT2, fill=FILL_IN)
        put(hs, r, 3, "not used by any formula: the 2027 quarterly margins run off the 2026 spread instead, "
                      "so 3Q27 stays continuous with the 3Q26 card margin", font=SMALL)
        anch[f"seas margin {qn}"] = f"History!$B${r}"; r += 1
    for qn in ["1Q", "2Q", "3Q", "4Q"]:
        put(hs, r, 1, f"Seasonal share of full-year nights / revenue, {qn}")
        put(hs, r, 2, DM.SEAS_SHARE[qn][0] / 100.0, fmt=F_PCT2, fill=FILL_IN)
        put(hs, r, 3, f"2023-2025 mean; revenue share {DM.SEAS_SHARE[qn][1]:.1f}%", font=SMALL)
        anch[f"seas nights {qn}"] = f"History!$B${r}"; r += 1
    for aq in ("1Q26", "2Q26"):
        put(hs, r, 1, f"{aq} adj. EBITDA margin (actual)")
        put(hs, r, 2, f"={col(ci['adj_ebitda_margin_pct'])}{qrow[aq]}/100", fmt=F_PCT2)
        put(hs, r, 3, "actual; used as the 2027 seasonal template", font=SMALL)
        anch[f"{aq} margin"] = f"History!$B${r}"; r += 1
    for lab, v, note in [("LTM revenue ($M)", DM.LTM_REV, "3Q25-2Q26"),
                         ("LTM adj. EBITDA ($M)", DM.LTM_EBITDA, "3Q25-2Q26"),
                         ("LTM FCF ($M)", DM.LTM_FCF, "3Q25-2Q26"),
                         ("LTM SBC ($M)", DM.LTM_SBC, "3Q25-2Q26")]:
        put(hs, r, 1, lab); put(hs, r, 2, v, fmt=F_NUM1, fill=FILL_IN); put(hs, r, 3, note, font=SMALL)
        anch[lab] = f"History!$B${r}"; r += 1

    # WS17 audit fix: these eleven FY2025 actuals used to sit as bare literals inside the FY2025A
    # memo formulas on Costs and Cash ("=161", "=-232", "=3789"...), with no source anywhere. They
    # are anchors like everything else above, so they belong here, sourced, and referenced.
    r += 1
    r = section(hs, r, "FY2025 CASH-BRIDGE ANCHORS - the FY2025A memo column on Costs and Cash. "
                       "No forecast year depends on these.", 8)
    put(hs, r, 1, "Item", font=H1, fill=FILL_H); put(hs, r, 2, "Value", font=H1, fill=FILL_H)
    put(hs, r, 3, "Source / formula", font=H1, fill=FILL_H); r += 1
    BR = "data/processed/abnb_fcf_bridge.csv, sum of 1Q25-4Q25"
    for lab, v, note in [
        ("FY2025 D&A ($M)", 91.0, BR + " (da)"),
        ("FY2025 D&A and other add-backs ($M)", 161.0, BR + " (da 91 + other_addbacks 70)"),
        ("FY2025 interest income ($M)", 705.0, BR + " (interest_income)"),
        ("FY2025 interest expense ($M)", 0.0, BR + " reports zero in every quarter of 2025; the "
                                                  "note coupon sits inside other income/(expense) "
                                                  "in that bridge. Memo only - the forecast years "
                                                  "take interest expense from Inputs"),
        ("FY2025 other income / (expense) ($M)", -112.0, BR + " (other_income_expense)"),
        ("FY2025 cash taxes ($M)", -232.0, "WS07: FY2025 cash taxes 1.9% of revenue "
                                           "(0.019 x $12,241M = $232M); the book tax provision was $626M"),
        ("FY2025 change in unearned fees ($M)", 127.0, BR + " (change_unearned_fees)"),
        ("FY2025 working-capital residual ($M)", -139.0,
         "WS07: FY24 and FY25 both about -$140M. With the eight lines above this reproduces the "
         "reported FY2025 FCF of $4,613M exactly"),
        ("FY2025 capex ($M)", -33.0, BR + " (capex)"),
    ]:
        put(hs, r, 1, lab); put(hs, r, 2, v, fmt=F_NUM1, fill=FILL_IN); put(hs, r, 3, note, font=SMALL)
        anch[lab] = f"History!$B${r}"; r += 1
    for lab, cix, note in [("FY2025 buybacks ($M)", "buybacks_musd", "sum of the 2025 quarters above"),
                           ("FY2025 RSU tax withholding ($M)", "rsu_tax_withholding_musd",
                            "sum of the 2025 quarters above")]:
        cl = col(ci[cix])
        put(hs, r, 1, lab)
        put(hs, r, 2, "=" + "+".join(f"{cl}{qrow[q]}" for q in fy25q), fmt=F_NUM1)
        put(hs, r, 3, note, font=SMALL)
        anch[lab] = f"History!$B${r}"; r += 1

    HA = lambda k: anch[k]

    # ------------------------------------------------------------------ REVENUE
    rs = wb.create_sheet("Revenue")
    rs.column_dimensions["A"].width = 48
    rs.column_dimensions["B"].width = 9
    for i in range(3, 13):
        rs.column_dimensions[col(i)].width = 13
    CQ = {q: col(3 + i) for i, q in enumerate(FQ)}          # C..H
    CFY = {2026: "I", 2027: "J", 2028: "K"}
    rr = 1
    put(rs, rr, 1, "Revenue build - quarterly 3Q26 to 4Q27 and annual FY2026E to FY2028E", font=TITLE); rr += 1
    put(rs, rr, 1, "nights by region x ADR (ex-FX x FX) = GBV; revenue = GBV x take rate x the FX timing "
                   "wedge, plus new business outside GBV. FY2026 = 1H26 actual + 3Q26 + 4Q26.", font=SMALL); rr += 2

    # The Costs sheet is written after Revenue but Revenue's margin rows reference it, so the Costs
    # row layout is fixed up front and reused when that sheet is built.
    CC = {2025: "C", 2026: "D", 2027: "E", 2028: "F"}
    COST_KEYS = ["rev", "gbv", "nights", "cor", "ops", "pd", "bpm", "fop", "ga", "nbc", "ai",
                 "cash_costs", "addb", "adj_unc", "margin_unc", "adj", "margin", "sbc", "sbc_pct",
                 "da", "opinc", "opmargin", "ii", "ie", "pretax", "taxrate", "ni"]
    _stride = len(COST_KEYS) + 3                             # rows + section banner + header + spacer
    COST_ROWS = {_s: {k: 6 + DM.SCENARIOS.index(_s) * _stride + i for i, k in enumerate(COST_KEYS)}
                 for _s in DM.SCENARIOS}

    REV = {}                                                 # (scen, row_key) -> row number
    for s in DM.SCENARIOS:
        rr = section(rs, rr, f"{s.upper()} CASE")
        put(rs, rr, 1, "Line", font=H1, fill=FILL_H)
        put(rs, rr, 2, "Unit", font=H1, fill=FILL_H)
        for i, q in enumerate(FQ):
            put(rs, rr, 3 + i, q, font=H1, fill=FILL_H)
        for i, y in enumerate(DM.YEARS):
            put(rs, rr, 9 + i, f"FY{y}E", font=H1, fill=FILL_H)
        rr += 1
        row = {}
        spec = []

        def line(key, label, unit, qform, fyform, fmt=F_NUM1, fill=None):
            """Two-pass: register the row number now, evaluate the formulas after every key exists."""
            nonlocal rr
            row[key] = rr
            spec.append((rr, label, unit, qform, fyform, fmt, fill))
            rr += 1

        def flush():
            for rw, label, unit, qform, fyform, fmt, fill in spec:
                put(rs, rw, 1, label)
                put(rs, rw, 2, unit)
                for q in FQ:
                    f = qform(q)
                    if f is not None:
                        put(rs, rw, 3 + FQ.index(q), f, fmt=fmt, fill=fill)
                for y in DM.YEARS:
                    f = fyform(y)
                    if f is not None:
                        put(rs, rw, 9 + DM.YEARS.index(y), f, fmt=fmt, fill=fill)

        # --- regional nights
        for rg in REGIONS:
            line(f"pn_{rg}", f"  Prior-year nights, {DM.RLABEL[rg]}", "M",
                 lambda q, rg=rg: f"={HA(DM.PRIOR_Q[q] + ' nights (M)')}*{HA(DM.PRIOR_Q[q] + ' ' + rg + ' share')}"
                 if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26")
                 else f"={CQ[DM.PRIOR_Q[q]]}{row['n_' + rg]}",
                 lambda y: None)
        for rg in REGIONS:
            line(f"g_{rg}", f"  Nights growth, {DM.RLABEL[rg]}", "y/y",
                 lambda q, rg=rg: f"={QIN(q, rg, s)}/100", lambda y: None, fmt=F_PCT1, fill=FILL_IN)
        for rg in REGIONS:
            line(f"d_{rg}", f"  Regulatory drag, {DM.RLABEL[rg]}", "pp",
                 lambda q, rg=rg: f"={ANN(DM.QYEAR[q], 'reg_mult', s)}*{RDG(DM.QYEAR[q], rg)}/100",
                 lambda y, rg=rg: f"={ANN(y, 'reg_mult', s)}*{RDG(y, rg)}/100" if y == 2028 else None,
                 fmt=F_PCT2)
        for rg in REGIONS:
            line(f"n_{rg}", f"  Nights, {DM.RLABEL[rg]}", "M",
                 lambda q, rg=rg: f"={CQ[q]}{row['pn_' + rg]}*(1+{CQ[q]}{row['g_' + rg]}-{CQ[q]}{row['d_' + rg]})",
                 lambda y, rg=rg: (f"={CQ['4Q26']}{row['n_' + rg]}/{CQ['4Q26']}{row['nights']}*{CFY[2026]}{row['nights']}" if y == 2026 else
                                   (f"=SUM({CQ['1Q27']}{row['n_' + rg]}:{CQ['4Q27']}{row['n_' + rg]})" if y == 2027 else
                                    f"={CFY[2027]}{row['n_' + rg]}/{CFY[2027]}{row['nights']}*{CFY[2028]}{row['nights']}")))
        line("nights", "Nights & experiences booked", "M",
             lambda q: "=" + "+".join(f"{CQ[q]}{row['n_' + rg]}" for rg in REGIONS),
             lambda y: (f"={HA('1H2026 nights (M)')}+{CQ['3Q26']}{row['nights']}+{CQ['4Q26']}{row['nights']}" if y == 2026 else
                        (f"=SUM({CQ['1Q27']}{row['nights']}:{CQ['4Q27']}{row['nights']})" if y == 2027 else
                         f"={CFY[2027]}{row['nights']}*(1+{ANN(2028, 'nights_2028', s)}-"
                         + "-".join(f"{CFY[2027]}{row['n_' + rg]}/{CFY[2027]}{row['nights']}*{CFY[2028]}{row['d_' + rg]}" for rg in REGIONS) + ")")),
             fmt=F_NUM1, fill=FILL_OUT)
        line("nights_yoy", "  Nights growth", "y/y",
             lambda q: f"={CQ[q]}{row['nights']}/" + (f"{HA(DM.PRIOR_Q[q] + ' nights (M)')}-1" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['nights']}-1"),
             lambda y: (f"={CFY[y]}{row['nights']}/{HA('FY2025 nights (M)')}-1" if y == 2026
                        else f"={CFY[y]}{row['nights']}/{CFY[y - 1]}{row['nights']}-1"), fmt=F_PCT1)
        # --- price
        line("adr_exfx", "ADR ex-FX growth", "y/y", lambda q: f"={QIN(q, 'adr_exfx', s)}/100",
             lambda y: f"={ANN(2028, 'adr_exfx_2028', s)}" if y == 2028 else None, fmt=F_PCT1, fill=FILL_IN)
        line("adr_fx", "FX effect on ADR", "pp", lambda q: f"={QIN(q, 'adr_fx', s)}/100",
             lambda y: f"={ANN(2028, 'adr_fx_2028', s)}" if y == 2028 else None, fmt=F_PCT2, fill=FILL_IN)
        line("adr", "ADR", "$",
             lambda q: f"=" + (f"{HA(DM.PRIOR_Q[q] + ' ADR ($)')}" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['adr']}")
                       + f"*(1+{CQ[q]}{row['adr_exfx']})*(1+{CQ[q]}{row['adr_fx']})",
             lambda y: f"={CFY[y]}{row['gbv']}/{CFY[y]}{row['nights']}", fmt=F_NUM2)
        line("adr_yoy", "  ADR growth (reported)", "y/y",
             lambda q: f"={CQ[q]}{row['adr']}/" + (f"{HA(DM.PRIOR_Q[q] + ' ADR ($)')}-1" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['adr']}-1"),
             lambda y: (f"={CFY[y]}{row['adr']}/{HA('FY2025 ADR ($)')}-1" if y == 2026
                        else f"={CFY[y]}{row['adr']}/{CFY[y - 1]}{row['adr']}-1"), fmt=F_PCT1)
        line("gbv", "Gross booking value", "$M",
             lambda q: f"={CQ[q]}{row['nights']}*{CQ[q]}{row['adr']}",
             lambda y: (f"={HA('1H2026 GBV ($M)')}+{CQ['3Q26']}{row['gbv']}+{CQ['4Q26']}{row['gbv']}" if y == 2026 else
                        (f"=SUM({CQ['1Q27']}{row['gbv']}:{CQ['4Q27']}{row['gbv']})" if y == 2027 else
                         f"={CFY[2028]}{row['nights']}*{CFY[2027]}{row['adr']}*(1+{CFY[2028]}{row['adr_exfx']})*(1+{CFY[2028]}{row['adr_fx']})")),
             fmt=F_NUM)
        line("gbv_yoy", "  GBV growth", "y/y",
             lambda q: f"={CQ[q]}{row['gbv']}/" + (f"{HA(DM.PRIOR_Q[q] + ' GBV ($M)')}-1" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['gbv']}-1"),
             lambda y: (f"={CFY[y]}{row['gbv']}/{HA('FY2025 GBV ($M)')}-1" if y == 2026
                        else f"={CFY[y]}{row['gbv']}/{CFY[y - 1]}{row['gbv']}-1"), fmt=F_PCT1)
        line("take", "Take rate, assumed (prior-year quarter + bps)", "%",
             lambda q: f"=" + (f"{HA(DM.PRIOR_Q[q] + ' take rate')}" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['take']}")
                       + f"+{ANN(DM.QYEAR[q], 'take_bps', s)}/10000",
             lambda y: ((f"=({HA('1H2026 revenue ($M)')}+{CQ['3Q26']}{row['gbv']}*{CQ['3Q26']}{row['take']}"
                         f"+{CQ['4Q26']}{row['gbv']}*{CQ['4Q26']}{row['take']})/{CFY[2026]}{row['gbv']}") if y == 2026 else
                        (("=(" + "+".join(f"{CQ[x]}{row['gbv']}*{CQ[x]}{row['take']}" for x in ["1Q27", "2Q27", "3Q27", "4Q27"])
                          + f")/{CFY[2027]}{row['gbv']}") if y == 2027 else
                         f"={CFY[2027]}{row['take']}+{ANN(2028, 'take_bps', s)}/10000")), fmt=F_PCT2)
        line("rev_fx", "FX effect on revenue (lagged)", "pp", lambda q: f"={QIN(q, 'rev_fx', s)}/100",
             lambda y: f"={ANN(2028, 'rev_fx_2028', s)}" if y == 2028 else None, fmt=F_PCT2, fill=FILL_IN)
        line("wedge", "  FX timing wedge (revenue FX less ADR FX)", "pp",
             lambda q: f"=(1+{CQ[q]}{row['rev_fx']})/(1+{CQ[q]}{row['adr_fx']})-1",
             lambda y: f"=(1+{CFY[2028]}{row['rev_fx']})/(1+{CFY[2028]}{row['adr_fx']})-1" if y == 2028 else None,
             fmt=F_PCT2)
        line("core", "Core revenue (GBV x take x wedge)", "$M",
             lambda q: f"={CQ[q]}{row['gbv']}*{CQ[q]}{row['take']}*(1+{CQ[q]}{row['wedge']})",
             lambda y: (f"={HA('1H2026 revenue ($M)')}+{CQ['3Q26']}{row['core']}+{CQ['4Q26']}{row['core']}" if y == 2026 else
                        (f"=SUM({CQ['1Q27']}{row['core']}:{CQ['4Q27']}{row['core']})" if y == 2027 else
                         f"={CFY[2028]}{row['gbv']}*{CFY[2028]}{row['take']}*(1+{CFY[2028]}{row['wedge']})")),
             fmt=F_NUM)
        line("take_impl", "  Implied reported take rate (core revenue / GBV)", "%",
             lambda q: f"={CQ[q]}{row['core']}/{CQ[q]}{row['gbv']}",
             lambda y: f"={CFY[y]}{row['core']}/{CFY[y]}{row['gbv']}", fmt=F_PCT2)
        line("nb", "New business outside GBV x take, incremental", "$M",
             lambda q: (f"={CFY[2026]}{row['nb']}*{CQ[q]}{row['core']}/({CQ['3Q26']}{row['core']}+{CQ['4Q26']}{row['core']})"
                        if q in ("3Q26", "4Q26") else
                        f"={CFY[2027]}{row['nb']}*{CQ[q]}{row['core']}/SUM({CQ['1Q27']}{row['core']}:{CQ['4Q27']}{row['core']})"),
             lambda y: f"={ANN(y, 'new_business', s)}-{HA('FY2025 new-business revenue outside GBV x take ($M)')}"
                       f"*{CFY[y]}{row['core']}/{HA('FY2025 revenue ($M)')}", fmt=F_NUM1)
        line("revenue", "REVENUE", "$M",
             lambda q: f"={CQ[q]}{row['core']}+{CQ[q]}{row['nb']}",
             lambda y: f"={CFY[y]}{row['core']}+{CFY[y]}{row['nb']}", fmt=F_NUM, fill=FILL_OUT)
        line("rev_yoy", "  Revenue growth", "y/y",
             lambda q: f"={CQ[q]}{row['revenue']}/" + (f"{HA(DM.PRIOR_Q[q] + ' revenue ($M)')}-1" if DM.PRIOR_Q[q] in ("3Q25", "4Q25", "1Q26", "2Q26") else f"{CQ[DM.PRIOR_Q[q]]}{row['revenue']}-1"),
             lambda y: (f"={CFY[y]}{row['revenue']}/{HA('FY2025 revenue ($M)')}-1" if y == 2026
                        else f"={CFY[y]}{row['revenue']}/{CFY[y - 1]}{row['revenue']}-1"),
             fmt=F_PCT1, fill=FILL_OUT)
        line("q_spread", "  Seasonal margin spread vs FY2026 (2027 quarters)", "pts",
             lambda q: (None if q in ("3Q26", "4Q26") else
                        ((f"={HA(DM.PRIOR_Q[q] + ' margin')}-Costs!${CC[2026]}${COST_ROWS[s]['margin']}")
                         if DM.PRIOR_Q[q] in ("1Q26", "2Q26") else
                         f"={QIN(DM.PRIOR_Q[q], 'margin', s)}/100-Costs!${CC[2026]}${COST_ROWS[s]['margin']}")),
             lambda y: None, fmt=F_PCT2)
        line("q_delta", "  Correction so the 2027 quarters reproduce the FY2027 margin", "pts",
             lambda q: (None if q in ("3Q26", "4Q26") else
                        "=-(" + "+".join(f"{CQ[x]}{row['q_spread']}*{CQ[x]}{row['revenue']}" for x in ["1Q27", "2Q27", "3Q27", "4Q27"])
                        + ")/SUM(" + f"{CQ['1Q27']}{row['revenue']}:{CQ['4Q27']}{row['revenue']}" + ")"),
             lambda y: None, fmt=F_PCT2)
        line("q_margin", "Adj. EBITDA margin", "%",
             lambda q: (f"={QIN(q, 'margin', s)}/100" if q in ("3Q26", "4Q26") else
                        f"=Costs!${CC[2027]}${COST_ROWS[s]['margin']}+{CQ[q]}{row['q_spread']}+{CQ[q]}{row['q_delta']}"),
             lambda y: f"=Costs!${CC[y]}${COST_ROWS[s]['margin']}", fmt=F_PCT1, fill=FILL_OUT)
        line("q_ebitda", "Adj. EBITDA", "$M",
             lambda q: f"={CQ[q]}{row['revenue']}*{CQ[q]}{row['q_margin']}",
             lambda y: f"=Costs!${CC[y]}${COST_ROWS[s]['adj']}", fmt=F_NUM, fill=FILL_OUT)
        line("seas_chk", "  Memo: seasonal share of full-year nights", "%",
             lambda q: (f"={CQ[q]}{row['nights']}/{CFY[2026]}{row['nights']}" if q in ("3Q26", "4Q26")
                        else f"={CQ[q]}{row['nights']}/{CFY[2027]}{row['nights']}"),
             lambda y: None, fmt=F_PCT1)
        line("seas_hist", "  Memo: 2023-2025 mean seasonal share", "%",
             lambda q: f"={HA('seas nights ' + q[:2])}", lambda y: None, fmt=F_PCT1)
        flush()
        REV[s] = dict(row)
        rr += 1

    # ------------------------------------------------------------------ COSTS
    cs = wb.create_sheet("Costs")
    cs.column_dimensions["A"].width = 52
    cs.column_dimensions["B"].width = 9
    for i in range(3, 8):
        cs.column_dimensions[col(i)].width = 14
    CC = {2025: "C", 2026: "D", 2027: "E", 2028: "F"}
    cr = 1
    put(cs, cr, 1, "Cost stack, adj. EBITDA and net income - lever by lever", font=TITLE); cr += 1
    put(cs, cr, 1, "Mechanics from WS07 07_margin_lever_model.py: cost of revenue on GBV dollars, "
                   "operations & support on nights, the rest on cash growth rates.", font=SMALL); cr += 2
    COST = {}
    for s in DM.SCENARIOS:
        cr = section(cs, cr, f"{s.upper()} CASE", 7)
        put(cs, cr, 1, "Line", font=H1, fill=FILL_H); put(cs, cr, 2, "Unit", font=H1, fill=FILL_H)
        for y in [2025, 2026, 2027, 2028]:
            put(cs, cr, 3 + [2025, 2026, 2027, 2028].index(y),
                f"FY{y}" + ("A" if y == 2025 else "E"), font=H1, fill=FILL_H)
        cr += 1
        row = {}
        cspec = []

        def cline(key, label, unit, f25, fy, fmt=F_NUM, fill=None):
            nonlocal cr
            assert COST_ROWS[s][key] == cr, (s, key, COST_ROWS[s][key], cr)
            row[key] = cr
            cspec.append((cr, label, unit, f25, fy, fmt, fill))
            cr += 1

        def cflush():
            for rw, label, unit, f25, fy, fmt, fill in cspec:
                put(cs, rw, 1, label); put(cs, rw, 2, unit)
                if f25 is not None:
                    put(cs, rw, 3, f25, fmt=fmt)
                for y in DM.YEARS:
                    f = fy(y)
                    if f is not None:
                        put(cs, rw, 3 + [2025, 2026, 2027, 2028].index(y), f, fmt=fmt, fill=fill)

        R = REV[s]
        rev_ref = lambda y: f"Revenue!${CFY[y]}${R['revenue']}"
        gbv_ref = lambda y: f"Revenue!${CFY[y]}${R['gbv']}"
        n_ref = lambda y: f"Revenue!${CFY[y]}${R['nights']}"
        cline("rev", "Revenue", "$M", f"={HA('FY2025 revenue ($M)')}", lambda y: f"={rev_ref(y)}")
        cline("gbv", "GBV", "$M", f"={HA('FY2025 GBV ($M)')}", lambda y: f"={gbv_ref(y)}")
        cline("nights", "Nights", "M", f"={HA('FY2025 nights (M)')}", lambda y: f"={n_ref(y)}", fmt=F_NUM1)
        cline("cor", "Cost of revenue, cash", "$M", f"={HA('FY2025 cost of revenue, cash ($M)')}",
              lambda y: f"={CC[y - 1]}{row['cor']}/{CC[y - 1]}{row['gbv']}*(1+{ANN(y, 'cor_per_gbv', s)})*{CC[y]}{row['gbv']}")
        cline("ops", "Operations & support, cash", "$M", f"={HA('FY2025 ops & support, cash ($M)')}",
              lambda y: f"={CC[y - 1]}{row['ops']}/{CC[y - 1]}{row['nights']}*(1+{ANN(y, 'ops_cpn', s)})*{CC[y]}{row['nights']}")
        cline("pd", "Product development, cash (ex-SBC)", "$M", f"={HA('FY2025 product development, cash ($M)')}",
              lambda y: f"={CC[y - 1]}{row['pd']}*(1+{ANN(y, 'pd_cash', s)})")
        cline("bpm", "Brand & performance marketing", "$M", f"={HA('FY2025 brand & performance marketing ($M)')}",
              lambda y: f"={CC[y - 1]}{row['bpm']}*(1+{ANN(y, 'bpm_cash', s)})")
        cline("fop", "Field operations & policy, cash", "$M", f"={HA('FY2025 field operations & policy, cash ($M)')}",
              lambda y: f"={CC[y - 1]}{row['fop']}*(1+{ANN(y, 'fop_cash', s)})")
        cline("ga", "G&A, cash", "$M", f"={HA('FY2025 G&A, cash ($M)')}",
              lambda y: f"={CC[y - 1]}{row['ga']}*(1+{ANN(y, 'ga_cash', s)})")
        cline("nbc", "New-business cost of the incremental revenue", "$M", "=0",
              lambda y: f"=Revenue!${CFY[y]}${R['nb']}*(1-{VAL('nb_incr_margin', s)})", fmt=F_NUM1)
        cline("ai", "AI referral cost", "$M", "=0",
              lambda y: f"={CC[y]}{row['rev']}*{ANN(y, 'ai_referral_pct', s)}", fmt=F_NUM1)
        cline("cash_costs", "Total cash costs", "$M",
              "=" + "+".join(f"C{row[k]}" for k in ["cor", "ops", "pd", "bpm", "fop", "ga", "nbc", "ai"]),
              lambda y: "=" + "+".join(f"{CC[y]}{row[k]}" for k in ["cor", "ops", "pd", "bpm", "fop", "ga", "nbc", "ai"]))
        cline("addb", "D&A and other add-backs", "$M", f"={HA('FY2025 D&A and other add-backs ($M)')}",
              lambda y: f"={CC[y]}{row['rev']}*{ANN(y, 'addback_pct', s)}", fmt=F_NUM1)
        cline("adj_unc", "  Adj. EBITDA before the WS07 38% ceiling", "$M", f"={HA('FY2025 adj. EBITDA ($M)')}",
              lambda y: f"={CC[y]}{row['rev']}-{CC[y]}{row['cash_costs']}+{CC[y]}{row['addb']}")
        cline("margin_unc", "  Margin before the ceiling", "%", f"=C{row['adj_unc']}/C{row['rev']}",
              lambda y: f"={CC[y]}{row['adj_unc']}/{CC[y]}{row['rev']}", fmt=F_PCT1)
        cline("adj", "ADJUSTED EBITDA (capped at the WS07 ceiling)", "$M", f"={HA('FY2025 adj. EBITDA ($M)')}",
              lambda y: f"=MIN({CC[y]}{row['adj_unc']},{CC[y]}{row['rev']}*{VAL('margin_cap', s)})", fill=FILL_OUT)
        cline("margin", "  Adj. EBITDA margin", "%", f"=C{row['adj']}/C{row['rev']}",
              lambda y: f"={CC[y]}{row['adj']}/{CC[y]}{row['rev']}", fmt=F_PCT1, fill=FILL_OUT)
        cline("sbc", "Stock-based compensation", "$M", f"={HA('FY2025 SBC ($M)')}",
              lambda y: f"={CC[y - 1]}{row['sbc']}*(1+{ANN(y, 'sbc_growth', s)})")
        cline("sbc_pct", "  SBC, % of revenue", "%", f"=C{row['sbc']}/C{row['rev']}",
              lambda y: f"={CC[y]}{row['sbc']}/{CC[y]}{row['rev']}", fmt=F_PCT1)
        cline("da", "D&A", "$M", f"={HA('FY2025 D&A ($M)')}",
              lambda y: f"={CC[y]}{row['rev']}*{ANN(y, 'da_pct', s)}", fmt=F_NUM1)
        cline("opinc", "GAAP operating income", "$M", None,
              lambda y: f"={CC[y]}{row['adj']}-{CC[y]}{row['sbc']}-{CC[y]}{row['addb']}")
        cline("opmargin", "  GAAP operating margin", "%", None,
              lambda y: f"={CC[y]}{row['opinc']}/{CC[y]}{row['rev']}", fmt=F_PCT1)
        cline("ii", "Interest income", "$M", f"={HA('FY2025 interest income ($M)')}",
              lambda y: f"={ANN(y, 'int_income', s)}", fmt=F_NUM1)
        cline("ie", "Interest expense", "$M", f"={HA('FY2025 interest expense ($M)')}",
              lambda y: f"={ANN(y, 'int_expense', s)}", fmt=F_NUM1)
        cline("pretax", "Pre-tax income", "$M", None,
              lambda y: f"={CC[y]}{row['opinc']}+{CC[y]}{row['ii']}-{CC[y]}{row['ie']}")
        cline("taxrate", "  Effective tax rate", "%", None,
              lambda y: f"={ANN(y, 'eff_tax_rate', s)}", fmt=F_PCT1)
        cline("ni", "Net income", "$M", None,
              lambda y: f"={CC[y]}{row['pretax']}*(1-{CC[y]}{row['taxrate']})", fill=FILL_OUT)
        cflush()
        COST[s] = dict(row)
        cr += 1

    # ------------------------------------------------------------------ CASH
    ks = wb.create_sheet("Cash")
    ks.column_dimensions["A"].width = 52
    ks.column_dimensions["B"].width = 9
    for i in range(3, 8):
        ks.column_dimensions[col(i)].width = 14
    kr = 1
    put(ks, kr, 1, "Free cash flow, buybacks and the share-count path", font=TITLE); kr += 1
    put(ks, kr, 1, "FCF bridge from WS07. Share count: buybacks and RSU issuance at a price growing "
                   "5% a year, 35% of SBC withheld for tax. Both roll-forwards start from the same 2Q26 "
                   "actual (net cash $9,593M, 597M diluted shares) and both consume the SAME flows: for "
                   "FY2026 only the 2H26 buyback (FY26 less the 1H26 actual $2,139M) and only 2H26 SBC "
                   "(FY26 less the 1H26 actual $897M); FY2027 and FY2028 take the full year. WS17 finding 3 "
                   "(the FY2026 share roll applied the FULL-YEAR buyback and SBC to a 30 Jun 2026 count, "
                   "double-counting the 1H26 repurchase and understating FY2026E shares by 8.55M) was "
                   "fixed by WS18 here and in 13_driver_model.py. See research/notes/overnight/"
                   "18_corrections-applied.md.",
        font=SMALL); kr += 2
    CASH = {}
    for s in DM.SCENARIOS:
        kr = section(ks, kr, f"{s.upper()} CASE", 7)
        put(ks, kr, 1, "Line", font=H1, fill=FILL_H); put(ks, kr, 2, "Unit", font=H1, fill=FILL_H)
        for y in [2025, 2026, 2027, 2028]:
            put(ks, kr, 3 + [2025, 2026, 2027, 2028].index(y),
                f"FY{y}" + ("A" if y == 2025 else "E"), font=H1, fill=FILL_H)
        kr += 1
        row = {}
        kspec = []

        def kline(key, label, unit, f25, fy, fmt=F_NUM, fill=None):
            nonlocal kr
            row[key] = kr
            kspec.append((kr, label, unit, f25, fy, fmt, fill))
            kr += 1

        def kflush():
            for rw, label, unit, f25, fy, fmt, fill in kspec:
                put(ks, rw, 1, label); put(ks, rw, 2, unit)
                if f25 is not None:
                    put(ks, rw, 3, f25, fmt=fmt)
                for y in DM.YEARS:
                    f = fy(y)
                    if f is not None:
                        put(ks, rw, 3 + [2025, 2026, 2027, 2028].index(y), f, fmt=fmt, fill=fill)

        C = COST[s]
        cref = lambda k, y: f"Costs!${CC[y]}${C[k]}"
        kline("adj", "Adjusted EBITDA", "$M", f"={cref('adj', 2025)}".replace(CC[2025], "C"),
              lambda y: f"={cref('adj', y)}")
        kline("ii", "+ Interest income", "$M", f"={HA('FY2025 interest income ($M)')}",
              lambda y: f"={cref('ii', y)}", fmt=F_NUM1)
        kline("ie", "- Interest expense", "$M", f"=-{HA('FY2025 interest expense ($M)')}",
              lambda y: f"=-{cref('ie', y)}", fmt=F_NUM1)
        kline("other", "+ Other income / (expense)", "$M", f"={HA('FY2025 other income / (expense) ($M)')}",
              lambda y: "=0", fmt=F_NUM1)
        kline("tax", "- Cash taxes", "$M", f"={HA('FY2025 cash taxes ($M)')}",
              lambda y: f"=-{cref('rev', y)}*{ANN(y, 'cash_tax_pct', s)}", fmt=F_NUM1)
        kline("unearn", "+ Change in unearned fees", "$M", f"={HA('FY2025 change in unearned fees ($M)')}",
              lambda y: f"={cref('rev', y)}*{ANN(y, 'd_unearned_pct', s)}", fmt=F_NUM1)
        kline("wc", "+ Working-capital residual", "$M", f"={HA('FY2025 working-capital residual ($M)')}",
              lambda y: f"={cref('rev', y)}*{ANN(y, 'wc_resid_pct', s)}", fmt=F_NUM1)
        kline("capex", "- Capex", "$M", f"={HA('FY2025 capex ($M)')}",
              lambda y: f"=-{cref('rev', y)}*{ANN(y, 'capex_pct', s)}", fmt=F_NUM1)
        kline("fcf", "FREE CASH FLOW", "$M",
              "=" + "+".join(f"C{row[k]}" for k in ["adj", "ii", "ie", "other", "tax", "unearn", "wc", "capex"]),
              lambda y: "=" + "+".join(f"{CC[y]}{row[k]}" for k in ["adj", "ii", "ie", "other", "tax", "unearn", "wc", "capex"]),
              fill=FILL_OUT)
        kline("fcfm", "  FCF margin", "%", f"=C{row['fcf']}/{HA('FY2025 revenue ($M)')}",
              lambda y: f"={CC[y]}{row['fcf']}/{cref('rev', y)}", fmt=F_PCT1)
        kline("conv", "  FCF / adj. EBITDA", "%", f"=C{row['fcf']}/C{row['adj']}",
              lambda y: f"={CC[y]}{row['fcf']}/{CC[y]}{row['adj']}", fmt=F_PCT1)
        kline("sbc", "SBC", "$M", f"={HA('FY2025 SBC ($M)')}", lambda y: f"={cref('sbc', y)}")
        kline("sbcfcf", "SBC-ADJUSTED FCF", "$M", f"=C{row['fcf']}-C{row['sbc']}",
              lambda y: f"={CC[y]}{row['fcf']}-{CC[y]}{row['sbc']}", fill=FILL_OUT)
        kline("sbcfcfm", "  SBC-adjusted FCF margin", "%", None,
              lambda y: f"={CC[y]}{row['sbcfcf']}/{cref('rev', y)}", fmt=F_PCT1)
        kline("px", "Share price used for buybacks and issuance", "$", f"={VAL('price', s)}",
              lambda y: f"={VAL('price', s)}*(1+{VAL('price_growth', s)})^{y - 2026}", fmt=F_NUM2)
        kline("bb", "Buybacks", "$M", f"={HA('FY2025 buybacks ($M)')}",
              lambda y: f"={ANN(y, 'buybacks', s)}", fmt=F_NUM1)
        kline("wh", "RSU tax withholding", "$M", f"={HA('FY2025 RSU tax withholding ($M)')}",
              lambda y: f"={CC[y]}{row['sbc']}*{VAL('withholding_pct', s)}", fmt=F_NUM1)
        # WS18 (WS17 finding 3): FY2026 rolls the 2Q26 count on 2H26 flows only -- the same
        # (FY26 less 1H26 actual) deltas the net-cash line below uses. FY2027/28 take the full year.
        kline("shares", "Diluted shares, period end (FY2025A col = the 2Q26 actual)", "M",
              f"={VAL('shares_2q26', s)}",
              lambda y: (f"={CC[y - 1]}{row['shares']}"
                         f"-({CC[2026]}{row['bb']}-{HA('1H2026 buybacks ($M)')})/{CC[y]}{row['px']}"
                         f"+({CC[2026]}{row['sbc']}-{HA('1H2026 SBC ($M)')})/{CC[y]}{row['px']}"
                         f"*(1-{VAL('withholding_pct', s)})" if y == 2026 else
                         f"={CC[y - 1]}{row['shares']}-{CC[y]}{row['bb']}/{CC[y]}{row['px']}"
                         f"+{CC[y]}{row['sbc']}/{CC[y]}{row['px']}*(1-{VAL('withholding_pct', s)})"),
              fmt=F_NUM1, fill=FILL_OUT)
        kline("fcfps", "FCF per share", "$", None,
              lambda y: f"={CC[y]}{row['fcf']}/{CC[y]}{row['shares']}", fmt=F_NUM2, fill=FILL_OUT)
        kline("sbcfcfps", "SBC-adjusted FCF per share", "$", None,
              lambda y: f"={CC[y]}{row['sbcfcf']}/{CC[y]}{row['shares']}", fmt=F_NUM2, fill=FILL_OUT)
        kline("eps", "Earnings proxy per share (net income / shares)", "$", None,
              lambda y: f"={cref('ni', y)}/{CC[y]}{row['shares']}", fmt=F_NUM2, fill=FILL_OUT)
        kline("netcash", "Net cash ex float, period end (FY2025A col = the 30 Jun 2026 actual)", "$M",
              f"={VAL('net_cash_2q26', s)}",
              lambda y: (f"=C{row['netcash']}+({CC[2026]}{row['fcf']}-{HA('1H2026 FCF ($M)')})"
                         f"-({CC[2026]}{row['bb']}-{HA('1H2026 buybacks ($M)')})"
                         f"-({CC[2026]}{row['wh']}-{HA('1H2026 RSU withholding ($M)')})" if y == 2026 else
                         f"={CC[y - 1]}{row['netcash']}+{CC[y]}{row['fcf']}-{CC[y]}{row['bb']}-{CC[y]}{row['wh']}"))
        kflush()
        CASH[s] = dict(row)
        kr += 1

    # ------------------------------------------------------------------ VALUATION
    vs = wb.create_sheet("Valuation")
    vs.column_dimensions["A"].width = 50
    vs.column_dimensions["B"].width = 10
    for i in range(3, 10):
        vs.column_dimensions[col(i)].width = 14
    vr = 1
    put(vs, vr, 1, "Valuation - exit multiples, DCF, football field, reverse DCF", font=TITLE); vr += 1
    put(vs, vr, 1, "Exit multiples from WS12; cost of equity from WS09/WS12. Every price is a live "
                   "formula off Costs and Cash.", font=SMALL); vr += 2

    # WS17 audit fix: column 5 on every scenario block below is an Active column driven by the
    # Inputs!B4 selector, so the selector (and the `Scenario` named range) actually does something.
    ACT = lambda rw: f"=CHOOSE({SCEN_SEL},B{rw},C{rw},D{rw})"
    vr = section(vs, vr, "1. FY2027E BASIS", 7)
    put(vs, vr, 1, "Item", font=H1, fill=FILL_H)
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, s, font=H1, fill=FILL_H)
    put(vs, vr, 5, "Active (Inputs!B4)", font=H1, fill=FILL_H)
    vr += 1
    basis = {}
    for key, label, ref, fmt in [
            ("ebitda", "Adj. EBITDA, FY2027E ($M)", lambda s: f"Costs!${CC[2027]}${COST[s]['adj']}", F_NUM),
            ("fcf", "FCF, FY2027E ($M)", lambda s: f"Cash!${CC[2027]}${CASH[s]['fcf']}", F_NUM),
            ("sbcfcf", "SBC-adjusted FCF, FY2027E ($M)", lambda s: f"Cash!${CC[2027]}${CASH[s]['sbcfcf']}", F_NUM),
            ("eps", "Earnings proxy per share, FY2027E ($)", lambda s: f"Cash!${CC[2027]}${CASH[s]['eps']}", F_NUM2),
            ("netcash", "Net cash ex float, FY2027E ($M)", lambda s: f"Cash!${CC[2027]}${CASH[s]['netcash']}", F_NUM),
            ("shares", "Diluted shares, FY2027E (M)", lambda s: f"Cash!${CC[2027]}${CASH[s]['shares']}", F_NUM1),
            ("ebitda28", "Adj. EBITDA, FY2028E ($M)", lambda s: f"Costs!${CC[2028]}${COST[s]['adj']}", F_NUM),
            ("netcash28", "Net cash ex float, FY2028E ($M)", lambda s: f"Cash!${CC[2028]}${CASH[s]['netcash']}", F_NUM),
            ("shares28", "Diluted shares, FY2028E (M)", lambda s: f"Cash!${CC[2028]}${CASH[s]['shares']}", F_NUM1),
            ("fcf28", "FCF, FY2028E ($M)", lambda s: f"Cash!${CC[2028]}${CASH[s]['fcf']}", F_NUM)]:
        put(vs, vr, 1, label)
        for i, s in enumerate(DM.SCENARIOS):
            put(vs, vr, 2 + i, f"={ref(s)}", fmt=fmt)
        put(vs, vr, 5, ACT(vr), fmt=fmt)
        basis[key] = vr
        vr += 1
    vr += 1

    vr = section(vs, vr, "2. PRICE PER LENS", 7)
    put(vs, vr, 1, "Lens", font=H1, fill=FILL_H)
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, s, font=H1, fill=FILL_H)
    put(vs, vr, 5, "Active (Inputs!B4)", font=H1, fill=FILL_H)
    put(vs, vr, 6, "Multiple source", font=H1, fill=FILL_H)
    vr += 1
    lensrow = {}
    B = lambda k, i: f"{col(2 + i)}{basis[k]}"
    for key, label, form, src in [
        ("ev_ebitda", "EV / adj. EBITDA, FY2027E",
         lambda i, s: f"=({VAL('exit_ev_ebitda', s)}*{B('ebitda', i)}+{B('netcash', i)})/{B('shares', i)}",
         "WS12 13.5 / 16.5 / 18.5x"),
        ("ev_fcf", "EV / FCF, FY2027E",
         lambda i, s: f"=({VAL('exit_ev_fcf', s)}*{B('fcf', i)}+{B('netcash', i)})/{B('shares', i)}",
         "5 Sep 15/19/23x rescaled by WS12's 0.75x haircut"),
        ("p_sbcfcf", "P / SBC-adjusted FCF, FY2027E",
         lambda i, s: f"={VAL('exit_p_sbcfcf', s)}*{B('sbcfcf', i)}/{B('shares', i)}",
         "same 0.75x haircut on 20/26/32x"),
        ("p_eps", "P / earnings proxy, FY2027E",
         lambda i, s: f"={VAL('exit_p_earnings', s)}*{B('eps', i)}", "same"),
        ("ev_ebitda28", "EV / adj. EBITDA, FY2028E",
         lambda i, s: f"=({VAL('exit_ev_ebitda', s)}*{B('ebitda28', i)}+{B('netcash28', i)})/{B('shares28', i)}",
         "WS12 multiples on the FY28 basis"),
    ]:
        put(vs, vr, 1, label)
        for i, s in enumerate(DM.SCENARIOS):
            put(vs, vr, 2 + i, form(i, s), fmt=F_NUM2, fill=FILL_OUT)
        put(vs, vr, 5, ACT(vr), fmt=F_NUM2, fill=FILL_OUT)
        put(vs, vr, 6, src, font=SMALL)
        lensrow[key] = vr
        vr += 1
    vr += 1

    vr = section(vs, vr, "3. DCF ON FY2027E FCF - 10-year fade to terminal growth, discounted at the cost of equity", 8)
    put(vs, vr, 1, "Year", font=H1, fill=FILL_H)
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, f"{s}: FCF ($M)", font=H1, fill=FILL_H)
        put(vs, vr, 5 + i, f"{s}: PV ($M)", font=H1, fill=FILL_H)
    vr += 1
    dcf_top = vr
    g0row = vr - 2
    put(vs, g0row, 8, "start growth = the 'DCF start growth' input, fading linearly to terminal growth", font=SMALL)
    for t in range(1, 11):
        put(vs, vr, 1, t)
        for i, s in enumerate(DM.SCENARIOS):
            g0 = f"{VAL('dcf_start_growth', s)}"
            gt = f"({g0}+({VAL('terminal_growth', s)}-{g0})*($A{vr}-1)/9)"
            prev = f"{col(2 + i)}{vr - 1}" if t > 1 else B('fcf', i)
            put(vs, vr, 2 + i, f"={prev}*(1+{gt})", fmt=F_NUM)
            put(vs, vr, 5 + i, f"={col(2 + i)}{vr}/(1+{VAL('cost_of_equity', s)})^$A{vr}", fmt=F_NUM)
        vr += 1
    put(vs, vr, 1, "Terminal value (Gordon), discounted")
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 5 + i, f"={col(2 + i)}{dcf_top + 9}*(1+{VAL('terminal_growth', s)})"
                           f"/({VAL('cost_of_equity', s)}-{VAL('terminal_growth', s)})"
                           f"/(1+{VAL('cost_of_equity', s)})^$A{dcf_top + 9}", fmt=F_NUM)
    tv_row = vr
    vr += 1
    put(vs, vr, 1, "Enterprise value ($M)")
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 5 + i, f"=SUM({col(5 + i)}{dcf_top}:{col(5 + i)}{dcf_top + 9})+{col(5 + i)}{tv_row}", fmt=F_NUM)
    ev_row = vr
    vr += 1
    put(vs, vr, 1, "DCF value per share ($)")
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, f"=({col(5 + i)}{ev_row}+{B('netcash', i)})/{B('shares', i)}", fmt=F_NUM2, fill=FILL_OUT)
    put(vs, vr, 8, "column E on this row is the Active (Inputs!B4) selection, not a PV", font=SMALL)
    put(vs, vr, 5, ACT(vr), fmt=F_NUM2, fill=FILL_OUT)
    lensrow["dcf"] = vr
    vr += 2

    vr = section(vs, vr, "4. FOOTBALL FIELD AND IMPLIED UPSIDE", 7)
    put(vs, vr, 1, "Measure", font=H1, fill=FILL_H)
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, s, font=H1, fill=FILL_H)
    put(vs, vr, 5, "Active (Inputs!B4)", font=H1, fill=FILL_H)
    vr += 1
    lens_keys = ["ev_ebitda", "ev_fcf", "p_sbcfcf", "p_eps", "ev_ebitda28", "dcf"]
    ff = {}
    for stat, fn in [("Low", "MIN"), ("High", "MAX"), ("Mean of lenses", "AVERAGE")]:
        put(vs, vr, 1, f"Football field: {stat}")
        for i, s in enumerate(DM.SCENARIOS):
            refs = ",".join(f"{col(2 + i)}{lensrow[k]}" for k in lens_keys)
            put(vs, vr, 2 + i, f"={fn}({refs})", fmt=F_NUM2, fill=FILL_OUT)
        put(vs, vr, 5, ACT(vr), fmt=F_NUM2, fill=FILL_OUT)
        ff[stat] = vr
        vr += 1
    put(vs, vr, 1, "Spot price ($)")
    for i, s in enumerate(DM.SCENARIOS):
        put(vs, vr, 2 + i, f"={VAL('price', s)}", fmt=F_NUM2)
    put(vs, vr, 5, ACT(vr), fmt=F_NUM2)
    spot_row = vr
    vr += 1
    for stat in ["Low", "High", "Mean of lenses"]:
        put(vs, vr, 1, f"Implied upside vs spot: {stat}")
        for i, s in enumerate(DM.SCENARIOS):
            put(vs, vr, 2 + i, f"={col(2 + i)}{ff[stat]}/{col(2 + i)}{spot_row}-1", fmt=F_PCT1, fill=FILL_OUT)
        put(vs, vr, 5, f"=E{ff[stat]}/E{spot_row}-1", fmt=F_PCT1, fill=FILL_OUT)
        vr += 1
    put(vs, vr, 1, "Memo: price on the 5 Sep exit multiples (18 / 22 / 25.5x)")
    for i, s in enumerate(DM.SCENARIOS):
        legacy = {"Bear": 18.0, "Base": 22.0, "Bull": 25.5}[s]
        put(vs, vr, 2 + i, f"=({legacy}*{B('ebitda', i)}+{B('netcash', i)})/{B('shares', i)}", fmt=F_NUM2)
    legacy_row = vr
    vr += 2

    vr = section(vs, vr, "5. REVERSE DCF - what growth the spot price discounts", 8)
    put(vs, vr, 1, "Assumed 10-year FCF growth", font=H1, fill=FILL_H)
    put(vs, vr, 2, "Value per share, reported FCF", font=H1, fill=FILL_H)
    put(vs, vr, 3, "Value per share, SBC-adjusted FCF", font=H1, fill=FILL_H)
    put(vs, vr, 5, "Constant growth for ten years then a Gordon terminal - the 5 Sep note's convention, so "
                   "the two are comparable. Reads off the base-case cost of equity and terminal growth; the "
                   "implied growth is where the value column crosses spot ($181.94).", font=SMALL)
    vr += 1
    rdcf_top = vr

    def rdcf(gref, j):
        """Value per share of a constant-growth ten-year stream plus a Gordon terminal, with the
        growth read from the cell `gref`. j = 0 reported FCF, j = 1 SBC-adjusted."""
        b = HA("LTM FCF ($M)") if j == 0 else f"({HA('LTM FCF ($M)')}-{HA('LTM SBC ($M)')})"
        coe, gt = VAL("cost_of_equity", "Base"), VAL("terminal_growth", "Base")
        # closed form of a linear fade is unwieldy in one cell; use a constant-growth annuity plus terminal
        f10 = f"{b}*(1+{gref})^10"
        pv = (f"{b}*(1+{gref})/({coe}-{gref})*(1-((1+{gref})/(1+{coe}))^10)"
              f"+{f10}*(1+{gt})/({coe}-{gt})/(1+{coe})^10")
        return f"=({pv}+{VAL('net_cash_2q26', 'Base')})/{VAL('shares_2q26', 'Base')}"

    for g in [0.03, 0.05, 0.07, 0.075, 0.09, 0.11, 0.13, 0.135, 0.15, 0.17]:
        put(vs, vr, 1, g, fmt=F_PCT1)
        for j in (0, 1):
            put(vs, vr, 2 + j, rdcf(f"A{vr}", j), fmt=F_NUM2)
        vr += 1
    put(vs, vr, 1, "Solved implied 10-year growth at spot (constant growth, then terminal):", font=H2)
    base_rev = [x for x in revdcf if x["cost_of_equity_pct"] == 10.5 and x["terminal_growth_pct"] == 3.0]
    for j, x in enumerate(base_rev):
        put(vs, vr, 2 + j, x["implied_10y_fcf_growth_pct"] / 100.0, fmt=F_PCT1, fill=FILL_IN)
        put(vs, vr, 5 + j, x["basis"], font=SMALL)
    solved_row = vr
    vr += 1
    # WS17 audit fix: the two cells above are an offline bisection, pasted in - they do not move when
    # the cost of equity or the terminal growth on Inputs changes. This row prices the same stream at
    # the solved growth and subtracts spot, so Excel itself says whether the paste is still current.
    put(vs, vr, 1, "  Check: value per share at the solved growth, less spot (0.00 = still current)",
        font=H2)
    for j in (0, 1):
        put(vs, vr, 2 + j,
            "=" + rdcf(f"{col(2 + j)}{solved_row}", j)[1:] + f"-{VAL('price', 'Base')}", fmt=F_NUM2)
    put(vs, vr, 5, "The solved cells above are yellow because they are an input, not a formula: a "
                   "bisection run in analysis/src/overnight/13_driver_model.py. If this row stops "
                   "reading 0.00, re-run that script - the cost of equity or the terminal growth moved.",
        font=SMALL)
    vr += 2

    vr = section(vs, vr, "6. SCENARIO GRID - FY2027E revenue growth x adj. EBITDA margin x exit multiple", 12)
    put(vs, vr, 1, "Price at the base exit multiple; rows = FY27 revenue growth, columns = FY27 margin",
        font=SMALL); vr += 1
    put(vs, vr, 1, "Revenue growth \\ margin", font=H1, fill=FILL_H)
    margins = [0.32, 0.34, 0.355, 0.365, 0.375, 0.39, 0.41]
    for i, m in enumerate(margins):
        put(vs, vr, 2 + i, m, font=H1, fill=FILL_IN, fmt=F_PCT1)
    grid_hdr = vr
    vr += 1
    fy26rev_base = f"Revenue!${CFY[2026]}${REV['Base']['revenue']}"
    nc = f"Cash!${CC[2027]}${CASH['Base']['netcash']}"
    sh = f"Cash!${CC[2027]}${CASH['Base']['shares']}"
    # WS17 audit fix: each cell used to restate its own row and column heading as literals, so the
    # axis labels and the arithmetic could drift apart. They now read the axis cells.
    for g in [0.04, 0.06, 0.08, 0.10, 0.12, 0.14, 0.16]:
        put(vs, vr, 1, g, fmt=F_PCT1, fill=FILL_IN)
        for i, m in enumerate(margins):
            put(vs, vr, 2 + i,
                f"=({VAL('exit_ev_ebitda', 'Base')}*{fy26rev_base}*(1+$A{vr})*{col(2 + i)}${grid_hdr}"
                f"+{nc})/{sh}", fmt=F_NUM2)
        vr += 1

    # ------------------------------------------------------------------ STREET
    ss = wb.create_sheet("Street")
    ss.column_dimensions["A"].width = 46
    for i in range(2, 7):
        ss.column_dimensions[col(i)].width = 16
    ss.column_dimensions["G"].width = 70
    sr = 1
    put(ss, sr, 1, "Street consensus versus our base case", font=TITLE); sr += 1
    put(ss, sr, 1, "Consensus from WS04 04_current_consensus.csv (Zacks and S&P Global Market "
                   "Intelligence, 3-4 Sep 2026). The bull case has to differ from these, not from LTM actuals.",
        font=SMALL); sr += 2
    put(ss, sr, 1, "Line", font=H1, fill=FILL_H)
    for i, t in enumerate(["Street", "n / range", "Our base", "Delta ($ or x)", "Delta (%)", "Vendor"]):
        put(ss, sr, 2 + i, t, font=H1, fill=FILL_H)
    sr += 1
    CB = COST["Base"]
    KB = CASH["Base"]
    RB = REV["Base"]
    street_rows = [
        ("FY2026 revenue ($M)", 14100.0, "8 est, 13,960-14,210", f"=Revenue!${CFY[2026]}${RB['revenue']}", "Zacks 4 Sep 26"),
        ("FY2026 revenue ($M), second source", 14160.0, "43 est, 13,800-14,300", f"=Revenue!${CFY[2026]}${RB['revenue']}", "S&P Global 3 Sep 26"),
        ("FY2026 adj. EPS ($)", 5.23, "13 est, 4.85-5.74", f"=Cash!${CC[2026]}${KB['eps']}", "Zacks 4 Sep 26"),
        ("FY2026 FCF ($M)", 5350.0, "-", f"=Cash!${CC[2026]}${KB['fcf']}", "S&P Global 3 Sep 26"),
        ("FY2027 revenue ($M)", 15730.0, "13 est, 14,990-16,290", f"=Revenue!${CFY[2027]}${RB['revenue']}", "Zacks 4 Sep 26"),
        ("FY2027 revenue ($M), second source", 15760.0, "43 est", f"=Revenue!${CFY[2027]}${RB['revenue']}", "S&P Global 3 Sep 26"),
        ("FY2027 adj. EPS ($)", 6.02, "13 est, 5.35-6.80", f"=Cash!${CC[2027]}${KB['eps']}", "Zacks 4 Sep 26"),
        ("3Q2026 revenue ($M)", 4740.0, "7 est, 4,720-4,770", f"=Revenue!${CQ['3Q26']}${RB['revenue']}", "Zacks 4 Sep 26"),
        ("3Q2026 adj. EPS ($)", 2.87, "11 est, 2.52-3.28", "", "Zacks 4 Sep 26"),
        ("4Q2026 revenue ($M)", 3200.0, "10 est, 3,050-3,700", f"=Revenue!${CQ['4Q26']}${RB['revenue']}", "Zacks 4 Sep 26"),
        ("Mean price target ($)", 178.96, "46 analysts, 125-220", f"=Valuation!$C${ff['Mean of lenses']}", "S&P Global / TipRanks 3 Sep 26"),
    ]
    for lab, st, rng, ours, vendor in street_rows:
        put(ss, sr, 1, lab)
        put(ss, sr, 2, st, fmt=F_NUM2 if st < 100 else F_NUM)
        put(ss, sr, 3, rng)
        if ours:
            put(ss, sr, 4, ours, fmt=F_NUM2 if st < 100 else F_NUM, fill=FILL_OUT)
            put(ss, sr, 5, f"=D{sr}-B{sr}", fmt=F_NUM2)
            put(ss, sr, 6, f"=D{sr}/B{sr}-1", fmt=F_PCT1)
        put(ss, sr, 7, vendor, font=SMALL)
        sr += 1
    sr += 1
    put(ss, sr, 1, "Historical beat behaviour (WS02, WS04) - how to read our base against the guide", font=H2); sr += 1
    for lab, v, note in [
        ("Median revenue beat vs guide midpoint, all 19 guided prints", 0.0252, "WS02 02_guidance_cushion_series.csv"),
        ("Median revenue beat vs guide midpoint, last 8 prints", 0.0179, "WS02; the cushion has halved"),
        ("Median revenue beat vs Street, post-2022 (n 14)", 0.0170, "WS04 04_consensus_at_print.csv"),
        ("Probability the print lands above the top of the guide range", 15 / 19, "WS02: 15 of 19"),
        ("Day-1 excess return explained by the beat", 0.04, "WS04 / the 5 Sep note: R-squared 0.04, LOO negative"),
        ("20-day drift zero-crossing, nights surprise vs Street", 0.0183, "WS04 post-2022 fit: excess_20d = -5.19 + 2.834 x nights surprise"),
    ]:
        put(ss, sr, 1, lab); put(ss, sr, 2, v, fmt=F_PCT2); put(ss, sr, 7, note, font=SMALL); sr += 1

    # ------------------------------------------------------------------ CARD 5 NOV
    cd = wb.create_sheet("Card_5Nov")
    cd.column_dimensions["A"].width = 44
    for i in range(2, 8):
        cd.column_dimensions[col(i)].width = 15
    cd.column_dimensions["H"].width = 76
    dr = 1
    put(cd, dr, 1, "5 November 2026 card - the Q3 2026 print", font=TITLE); dr += 1
    put(cd, dr, 1, "Model columns are live formulas off the Revenue and Costs sheets; the Active "
                   "column follows the scenario selector on Inputs!B4. External bars are constants "
                   "with their source. Q3 2026 reports 5 Nov 2026 (Zacks expected date).", font=SMALL); dr += 2
    put(cd, dr, 1, "Line", font=H1, fill=FILL_H)
    for i, t in enumerate(["Bear", "Base", "Bull", "Active (Inputs!B4)", "Company guide",
                           "Street / cushion bar", "Source of the bar"]):
        put(cd, dr, 2 + i, t, font=H1, fill=FILL_H)
    dr += 1

    def card(label, ref, guide, bar, src, fmt=F_NUM1):
        nonlocal dr
        put(cd, dr, 1, label)
        for i, s in enumerate(DM.SCENARIOS):
            if ref:
                put(cd, dr, 2 + i, ref(s), fmt=fmt, fill=(FILL_OUT if s == "Base" else None))
        if ref:
            # WS17 audit fix: the scenario selector now drives a real column on this sheet
            put(cd, dr, 5, f"=CHOOSE({SCEN_SEL},B{dr},C{dr},D{dr})", fmt=fmt, fill=FILL_OUT)
        put(cd, dr, 6, guide)
        put(cd, dr, 7, bar)
        put(cd, dr, 8, src, font=SMALL)
        dr += 1

    q3 = CQ["3Q26"]
    card("3Q26 revenue ($M)", lambda s: f"=Revenue!${q3}${REV[s]['revenue']}",
         "$4,690-4,770m (+15-17%)", "Street $4,740m; cushion-adj $4,815m",
         "2Q26 letter; WS04 Zacks 7 est; WS02 02_q3_2026_guide_card.csv (last-8 median cushion +1.79%)", F_NUM)
    card("  revenue growth y/y", lambda s: f"=Revenue!${q3}${REV[s]['rev_yoy']}",
         "+15% to +17%", "+15.79% (Street)", "WS02, WS04", F_PCT1)
    card("3Q26 nights (M)", lambda s: f"=Revenue!${q3}${REV[s]['nights']}",
         "low double-digit growth", "144-146m (derived)",
         "WS04 04_q3_2026_breakeven.csv; WS02 puts the cushion-adjusted range at 150-154m", F_NUM1)
    card("  nights growth y/y", lambda s: f"=Revenue!${q3}${REV[s]['nights_yoy']}",
         "'low double digit' (10-12%)", "WS08 nowcast +12.1% (9.3-14.9)",
         "WS08 08_q3_2026_nowcast.csv demand_eq; naive baseline +10.3%", F_PCT1)
    card("3Q26 ADR ($)", lambda s: f"=Revenue!${q3}${REV[s]['adr']}", "'up moderately'", "-", "2Q26 letter", F_NUM2)
    card("  ADR growth y/y", lambda s: f"=Revenue!${q3}${REV[s]['adr_yoy']}", "-", "+4.05% (WS08)",
         "WS08: ex-FX run-rate +3.25% plus FX +0.80pp; r 0.96, walk-forward 0.44x naive", F_PCT1)
    card("  of which FX", lambda s: f"=Revenue!${q3}${REV[s]['adr_fx']}", "-", "+0.80pp",
         "WS05 05_fx_schedule.csv broad-USD fit; WS08 refit", F_PCT2)
    card("3Q26 GBV ($M)", lambda s: f"=Revenue!${q3}${REV[s]['gbv']}", "mid-teens growth",
         "$26.6-27.2bn (WS02 cushion)", "2Q26 letter; WS02", F_NUM)
    card("  GBV growth y/y", lambda s: f"=Revenue!${q3}${REV[s]['gbv_yoy']}", "mid teens (14-16%)", "16-19%", "WS02", F_PCT1)
    card("3Q26 take rate", lambda s: f"=Revenue!${q3}${REV[s]['take']}", "'relatively in line y/y'",
         "17.9% +/- 0.2pt", "2Q26 call; WS02. 3Q25 actual 17.88%", F_PCT2)
    card("3Q26 adj. EBITDA margin", lambda s: f"=Revenue!${q3}${REV[s]['q_margin']}",
         "'down slightly' vs 50.1%", "48.7-50.0%",
         "WS07 5 Nov card (bear 48.3 / base 49.0 / bull 50.2); WS02 says 9 of 10 such ceilings were met", F_PCT1)
    card("3Q26 adj. EBITDA ($M)", lambda s: f"=Revenue!${q3}${REV[s]['q_ebitda']}",
         "up y/y", "$2,300-2,400m (derived)", "WS04 04_q3_2026_breakeven.csv", F_NUM)
    dr += 1
    put(cd, dr, 1, "THE Q4 GUIDE - what the print's reaction actually runs on", font=H2); dr += 1
    q4 = CQ["4Q26"]
    card("4Q26 revenue ($M), our forecast", lambda s: f"=Revenue!${q4}${REV[s]['revenue']}",
         "no guide yet", "Street $3,200m (3,050-3,700)",
         "WS04 Zacks 10 est - the widest spread in the table (21% high-low)", F_NUM)
    card("  4Q26 revenue growth", lambda s: f"=Revenue!${q4}${REV[s]['rev_yoy']}", "-", "+11-13% expected",
         "WS05: FX goes from about +3pp in Q3 to -0.4pp in Q4 and 84% of it is already realised", F_PCT1)
    card("  4Q26 nights growth", lambda s: f"=Revenue!${q4}${REV[s]['nights_yoy']}", "-", "accelerating vs Q3 is the bull tell",
         "WS05: post-2022 an accelerating nights guide averaged +6.7% day-1 (n 5), a decelerating one -5.4% (n 9)", F_PCT1)
    card("FY26 revenue ($M)", lambda s: f"=Revenue!${CFY[2026]}${REV[s]['revenue']}",
         "'at least mid teens' growth", "Street $14,100-14,160m", "2Q26 letter; WS04", F_NUM)
    card("  FY26 revenue growth", lambda s: f"=Revenue!${CFY[2026]}${REV[s]['rev_yoy']}",
         "at least +14-16%", "+15.2% to +15.6%", "WS04", F_PCT1)
    card("FY26 adj. EBITDA margin", lambda s: f"=Costs!${CC[2026]}${COST[s]['margin']}",
         "'at least 35.5%'", "36.1-36.9% on the FY24/FY25 cushion",
         "WS02: FY24 beat the floor by 140bp, FY25 by 60bp", F_PCT1)
    dr += 1
    put(cd, dr, 1, "CONTEXT AND BASE RATES", font=H2); dr += 1
    for lab, v, fmt, note in [
        ("Alt-data demand index, 3Q26 to date (z)", -0.325, F_NUM2,
         "WS08 08_demand_index_quarterly.csv, demand_eq; 2Q26 was -0.509, so the index improved. Only 4 of 7 components are populated for 3Q26 and no composite beat a naive baseline walk-forward."),
        ("Funds-held backlog nowcast, 3Q26 revenue growth", 0.120, F_PCT1,
         "WS08 08_backlog_tests.csv: funds held +10.46% y/y at 2Q26 x slope 0.60. Under-shot 1Q26 by 3.4pp and 2Q26 by 2.8pp because RNPL defers guest cash to check-in; add ~3pp."),
        ("Mean absolute day-1 move on a print", 0.0707, F_PCT1, "WS09, n 23; median 6.87%"),
        ("Mean 20-session post-print excess drift", -0.037, F_PCT1, "WS09, n 23, t -2.16"),
        ("Nights-acceleration sign rule hit rate", 17 / 21, F_PCT1,
         "WS01/predictive study: the sign of the nights-growth acceleration has set the day-1 direction in 17 of 21 prints"),
        ("Share of day-1 move that is re-rating, not estimates", 0.97, F_PCT2,
         "WS12: day-1 move correlates +0.97 with the change in the EV/NTM-revenue multiple, +0.09 with the estimate change"),
        ("One turn of EV / NTM EBITDA", 9.11, F_NUM2, "WS12: $9.11 per share = 5.0% at $181.94"),
        ("November excess return vs QQQ, 2021-2025", -0.051, F_PCT1, "WS09: ABNB has underperformed in every November since 2021"),
    ]:
        put(cd, dr, 1, lab)
        put(cd, dr, 3, v, fmt=fmt)
        put(cd, dr, 8, note, font=SMALL, wrap=True)
        dr += 1

    # ------------------------------------------------------------------ RECON
    rc = wb.create_sheet("Recon")
    rc.column_dimensions["A"].width = 46
    for i in range(2, 6):
        rc.column_dimensions[col(i)].width = 18
    rc.column_dimensions["F"].width = 40
    q = 1
    put(rc, q, 1, "Reconciliation: the Python mirror against this workbook", font=TITLE); q += 1
    put(rc, q, 1, "Column B is the Python mirror (13_driver_model.py), column C is this workbook and "
                  "column D is a live Excel formula, so Excel re-checks the difference every time the "
                  "file is opened. Every |delta| should be 0.00. The same 216 outputs are also checked "
                  "outside Excel by analysis/src/overnight/13_xlsx_eval.py. Workstream 17 then rebuilt "
                  "the file in Excel 16.0 itself (Application.CalculateFullRebuild via COM): 0 error "
                  "cells in 5,544, no circular reference, all 216 outputs and all 2,348 formula cells "
                  "agreeing with the mirror to better than 1e-6 relative (worst 4.2e-15). Full results: "
                  "data/processed/overnight/13_reconciliation.csv and 17_excel_vs_python.csv; the audit "
                  "is research/notes/overnight/17_excel-audit.md.",
        font=SMALL, wrap=True); q += 2
    put(rc, q, 1, "Item", font=H1, fill=FILL_H)
    for i, t in enumerate(["Python mirror", "Workbook cell", "Delta", "Cell reference"]):
        put(rc, q, 2 + i, t, font=H1, fill=FILL_H)
    q += 1
    checks = []
    for s in DM.SCENARIOS:
        for y in DM.YEARS:
            a = A[(s, y)]
            for lab, ref, val, fmt in [
                (f"{s} FY{y} revenue ($M)", f"Revenue!${CFY[y]}${REV[s]['revenue']}", a["revenue"], F_NUM),
                (f"{s} FY{y} nights (M)", f"Revenue!${CFY[y]}${REV[s]['nights']}", a["nights"], F_NUM1),
                (f"{s} FY{y} GBV ($M)", f"Revenue!${CFY[y]}${REV[s]['gbv']}", a["gbv"], F_NUM),
                (f"{s} FY{y} adj. EBITDA ($M)", f"Costs!${CC[y]}${COST[s]['adj']}", a["adj_ebitda"], F_NUM),
                (f"{s} FY{y} adj. EBITDA margin", f"Costs!${CC[y]}${COST[s]['margin']}", a["adj_ebitda_margin_pct"] / 100, F_PCT2),
                (f"{s} FY{y} net income ($M)", f"Costs!${CC[y]}${COST[s]['ni']}", a["net_income"], F_NUM),
                (f"{s} FY{y} FCF ($M)", f"Cash!${CC[y]}${CASH[s]['fcf']}", a["fcf"], F_NUM),
                (f"{s} FY{y} SBC-adj FCF ($M)", f"Cash!${CC[y]}${CASH[s]['sbcfcf']}", a["sbc_adj_fcf"], F_NUM),
                (f"{s} FY{y} diluted shares (M)", f"Cash!${CC[y]}${CASH[s]['shares']}", a["shares_end"], F_NUM1),
                (f"{s} FY{y} FCF per share ($)", f"Cash!${CC[y]}${CASH[s]['fcfps']}", a["fcf_per_share"], F_NUM2),
                (f"{s} FY{y} net cash ($M)", f"Cash!${CC[y]}${CASH[s]['netcash']}", a["net_cash"], F_NUM),
            ]:
                put(rc, q, 1, lab)
                put(rc, q, 2, val, fmt=fmt)
                put(rc, q, 3, f"={ref}", fmt=fmt)
                put(rc, q, 4, f"=C{q}-B{q}", fmt="0.0000")
                put(rc, q, 5, ref, font=SMALL)
                checks.append(dict(item=lab, python_value=round(val, 6), workbook_value=None,
                                   delta=None, status="pending", cell_reference=ref,
                                   method="workbook formula evaluated by analysis/src/overnight/13_xlsx_eval.py"))
                q += 1
        for qtr in FQ:
            qq = Q[(s, qtr)]
            for lab, ref, val, fmt in [
                (f"{s} {qtr} revenue ($M)", f"Revenue!${CQ[qtr]}${REV[s]['revenue']}", qq["revenue"], F_NUM),
                (f"{s} {qtr} nights (M)", f"Revenue!${CQ[qtr]}${REV[s]['nights']}", qq["nights"], F_NUM1),
                (f"{s} {qtr} ADR ($)", f"Revenue!${CQ[qtr]}${REV[s]['adr']}", qq["adr"], F_NUM2),
                (f"{s} {qtr} adj. EBITDA ($M)", f"Revenue!${CQ[qtr]}${REV[s]['q_ebitda']}", qq["adj_ebitda"], F_NUM),
                (f"{s} {qtr} adj. EBITDA margin", f"Revenue!${CQ[qtr]}${REV[s]['q_margin']}", qq["adj_ebitda_margin_pct"] / 100, F_PCT2),
            ]:
                put(rc, q, 1, lab); put(rc, q, 2, val, fmt=fmt); put(rc, q, 3, f"={ref}", fmt=fmt)
                put(rc, q, 4, f"=C{q}-B{q}", fmt="0.0000"); put(rc, q, 5, ref, font=SMALL)
                checks.append(dict(item=lab, python_value=round(val, 6), workbook_value=None,
                                   delta=None, status="pending", cell_reference=ref,
                                   method="workbook formula evaluated by analysis/src/overnight/13_xlsx_eval.py"))
                q += 1
        for stat, rw in [("Football field low", ff["Low"]), ("Football field high", ff["High"]),
                         ("Football field mean", ff["Mean of lenses"])]:
            val = [x for x in valrows if x["scenario"] == s and x["lens"] == stat][0]["price"]
            ref = f"Valuation!${col(2 + DM.SCENARIOS.index(s))}${rw}"
            put(rc, q, 1, f"{s} {stat}"); put(rc, q, 2, val, fmt=F_NUM2)
            put(rc, q, 3, f"={ref}", fmt=F_NUM2); put(rc, q, 4, f"=C{q}-B{q}", fmt="0.0000")
            put(rc, q, 5, ref, font=SMALL)
            checks.append(dict(item=f"{s} {stat}", python_value=round(val, 6), workbook_value=None,
                               delta=None, status="pending", cell_reference=ref,
                               method="workbook formula evaluated by analysis/src/overnight/13_xlsx_eval.py"))
            q += 1
        for lens, key in [("EV / adj. EBITDA, FY27E", "ev_ebitda"), ("EV / FCF, FY27E", "ev_fcf"),
                          ("P / SBC-adjusted FCF, FY27E", "p_sbcfcf"), ("P / earnings proxy, FY27E", "p_eps"),
                          ("EV / adj. EBITDA, FY28E", "ev_ebitda28"), ("DCF on FCF", "dcf")]:
            pyname = {"ev_ebitda": "EV / adj. EBITDA, FY27E", "ev_fcf": "EV / FCF, FY27E",
                      "p_sbcfcf": "P / SBC-adjusted FCF, FY27E", "p_eps": "P / earnings proxy, FY27E",
                      "ev_ebitda28": "EV / adj. EBITDA, FY28E", "dcf": "DCF on FCF"}[key]
            val = prices[s][pyname]
            ref = f"Valuation!${col(2 + DM.SCENARIOS.index(s))}${lensrow[key]}"
            put(rc, q, 1, f"{s} price: {lens}")
            put(rc, q, 2, val, fmt=F_NUM2); put(rc, q, 3, f"={ref}", fmt=F_NUM2)
            put(rc, q, 4, f"=C{q}-B{q}", fmt="0.0000"); put(rc, q, 5, ref, font=SMALL)
            checks.append(dict(item=f"{s} price: {lens}", python_value=round(val, 6),
                               workbook_value=None, delta=None, status="pending", cell_reference=ref,
                               method="workbook formula evaluated by analysis/src/overnight/13_xlsx_eval.py"))
            q += 1

    for sheet in wb.worksheets:
        sheet.freeze_panes = "B5"
    from openpyxl.workbook.defined_name import DefinedName
    for nm, ref in [("Scenario", SCEN_SEL.replace("Inputs!", "Inputs!")),
                    ("SharePrice", f"Inputs!$F${rv['price']}"),
                    ("CostOfEquity", f"Inputs!$F${rv['cost_of_equity']}"),
                    ("ExitMultiple", f"Inputs!$F${rv['exit_ev_ebitda']}")]:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))
    path = os.path.join(ROOT, "model", "ABNB_driver_model.xlsx")
    wb.save(path)

    # ---- evaluate the saved workbook and score every check for real
    _es = importlib.util.spec_from_file_location("xe13", os.path.join(HERE, "13_xlsx_eval.py"))
    XE = importlib.util.module_from_spec(_es)
    _es.loader.exec_module(XE)
    ev = XE.evaluate_workbook(path)
    for c in checks:
        ref = c["cell_reference"]
        sh, coord = ref.split("!")
        got = ev.cell(sh, coord.replace("$", ""))
        c["workbook_value"] = round(got, 6)
        c["delta"] = round(got - c["python_value"], 6)
        c["status"] = "ok" if abs(c["delta"]) <= max(1e-6, abs(c["python_value"]) * 1e-9) else "MISMATCH"
    return checks
