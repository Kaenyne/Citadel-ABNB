"""Workstream 17: audit `model/ABNB_driver_model.xlsx` against a real Excel recalculation.

READS
  data/processed/overnight/17_excel_recalc_dump.csv   every non-empty cell of the workbook after
                                                      Excel 16.0 CalculateFullRebuild (written by
                                                      the PowerShell COM driver, see the note)
  data/processed/overnight/13_reconciliation.csv      the 216 named outputs of workstream 13
  data/processed/overnight/13_valuation_summary.csv   bear/base/bull prices from the Python mirror
  data/processed/overnight/13_model_annual.csv        bear/base/bull annual lines
  model/ABNB_driver_model.xlsx                        re-evaluated by 13_xlsx_eval.py

WRITES
  data/processed/overnight/17_excel_vs_python.csv     216 outputs: Excel vs Python vs evaluator
  data/processed/overnight/17_all_formula_cells.csv   all 2,303 formula cells: Excel vs evaluator
  data/processed/overnight/17_formula_review.csv      static formula-review findings

RUN   py -3.13 analysis/src/overnight/17_excel_audit.py
"""
from __future__ import annotations

import csv
import importlib.util
import os
import re

import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
OD = lambda n: os.path.join(ROOT, "data", "processed", "overnight", n)   # noqa: E731
XLSX = os.path.join(ROOT, "model", "ABNB_driver_model.xlsx")

_es = importlib.util.spec_from_file_location("xe13", os.path.join(HERE, "13_xlsx_eval.py"))
XE = importlib.util.module_from_spec(_es)
_es.loader.exec_module(XE)

_dm = importlib.util.spec_from_file_location("dm13", os.path.join(HERE, "13_driver_model.py"))
DM = importlib.util.module_from_spec(_dm)
_dm.loader.exec_module(DM)

REF = re.compile(
    r"(?:(?:'(?P<sq>[^']+)'|(?P<sheet>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+))?")
FUNC = re.compile(r"\b([A-Z][A-Z0-9.]*)\s*\(")


def rel(a, b):
    if a is None or b is None:
        return None
    d = abs(a - b)
    return d / max(abs(a), abs(b), 1e-12)


# ---------------------------------------------------------------------------- load the Excel dump
def load_dump():
    d = pd.read_csv(OD("17_excel_recalc_dump.csv"))
    d["formula"] = d["formula"].fillna("")
    d["value"] = d["value"].fillna("")
    val, txt = {}, {}
    for r in d.itertuples():
        key = (r.sheet, r.address)
        txt[key] = str(r.value)
        try:
            val[key] = float(r.value)
        except (TypeError, ValueError):
            val[key] = None
    return d, val, txt


def a1_to_r1c1(formula, row, col):
    """Rewrite a formula's relative refs as offsets so two cells of the same row can be compared."""
    def sub(m):
        sh = m.group("sq") or m.group("sheet")
        pre = (sh + "!") if sh else ""
        out = []
        for c, r in ((m.group("c1"), m.group("r1")), (m.group("c2"), m.group("r2"))):
            if c is None:
                continue
            whole = m.group(0)
            abs_c = ("$" + c) in whole
            abs_r = ("$" + r) in whole
            ci = column_index_from_string(c)
            ri = int(r)
            cs = c if (abs_c or sh) else f"C[{ci - col:+d}]"
            rs = r if (abs_r or sh) else f"R[{ri - row:+d}]"
            out.append(f"{rs}{cs}")
        return pre + ":".join(out)
    return REF.sub(sub, formula)


def python_mirror():
    """Re-run the workstream-13 Python mirror in process, at full precision (the CSV it wrote
    rounds to four decimals, which is coarser than a 1e-6 relative test on margins and per-share
    numbers). Returns {recon item label: value} using the builder's own label conventions."""
    out, prices = {}, {}
    for scen in DM.SCENARIOS:
        ann = DM.build_annual(scen)
        qs = {q["quarter"]: q for q in DM.quarterly_pl(scen, ann)}
        A = {a["year"]: a for a in ann}
        for y in DM.YEARS:
            a = A[y]
            out[f"{scen} FY{y} revenue ($M)"] = a["revenue"]
            out[f"{scen} FY{y} nights (M)"] = a["nights"]
            out[f"{scen} FY{y} GBV ($M)"] = a["gbv"]
            out[f"{scen} FY{y} adj. EBITDA ($M)"] = a["adj_ebitda"]
            out[f"{scen} FY{y} adj. EBITDA margin"] = a["adj_ebitda_margin_pct"] / 100
            out[f"{scen} FY{y} net income ($M)"] = a["net_income"]
            out[f"{scen} FY{y} FCF ($M)"] = a["fcf"]
            out[f"{scen} FY{y} SBC-adj FCF ($M)"] = a["sbc_adj_fcf"]
            out[f"{scen} FY{y} diluted shares (M)"] = a["shares_end"]
            out[f"{scen} FY{y} FCF per share ($)"] = a["fcf_per_share"]
            out[f"{scen} FY{y} net cash ($M)"] = a["net_cash"]
        for qtr in DM.FQ:
            q = qs[qtr]
            out[f"{scen} {qtr} revenue ($M)"] = q["revenue"]
            out[f"{scen} {qtr} nights (M)"] = q["nights"]
            out[f"{scen} {qtr} ADR ($)"] = q["adr"]
            out[f"{scen} {qtr} adj. EBITDA ($M)"] = q["adj_ebitda"]
            out[f"{scen} {qtr} adj. EBITDA margin"] = q["adj_ebitda_margin_pct"] / 100
        v, px, _ = DM.valuation(ann, scen)
        prices[scen] = px
        for row in v:
            if row["lens"].startswith("Football field"):
                out[f"{scen} {row['lens']}"] = row["price"]
        for lens in ("EV / adj. EBITDA, FY27E", "EV / FCF, FY27E", "P / SBC-adjusted FCF, FY27E",
                     "P / earnings proxy, FY27E", "EV / adj. EBITDA, FY28E", "DCF on FCF"):
            out[f"{scen} price: {lens}"] = px[lens]
    return out, prices


def main():
    d, xval, xtxt = load_dump()
    pym, _prices = python_mirror()
    ev = XE.evaluate_workbook(XLSX)
    formulas = d[d.formula.str.startswith("=")]

    # ---------------------------------------------------------------- 1. all formula cells
    allrows, mism = [], 0
    for r in formulas.itertuples():
        try:
            got = ev.cell(r.sheet, r.address)
            err = ""
        except Exception as exc:                                            # noqa: BLE001
            got, err = None, type(exc).__name__ + ": " + str(exc)[:120]
        x = xval[(r.sheet, r.address)]
        rl = rel(x, got) if isinstance(got, (int, float)) else None
        ad = (None if (x is None or not isinstance(got, (int, float))) else abs(x - got))
        # a relative test is meaningless when both sides are ~0 (the Recon delta column), so an
        # absolute floor of 1e-6 (the workbook's unit is $M, or a fraction) counts as a match
        ok = (rl is not None and rl <= 1e-6) or (ad is not None and ad <= 1e-6)
        if not ok:
            mism += 1
        allrows.append(dict(sheet=r.sheet, address=r.address, formula=r.formula,
                            excel_value=x, evaluator_value=got, abs_diff=ad,
                            rel_diff=rl, pass_1e6=("yes" if ok else "no"), note=err))
    pd.DataFrame(allrows).to_csv(OD("17_all_formula_cells.csv"), index=False)
    print(f"[1] all formula cells: {len(allrows)} compared, {mism} outside 1e-6 relative")

    # ---------------------------------------------------------------- 2. the 216 named outputs
    recon = pd.read_csv(OD("13_reconciliation.csv"))
    rows, bad = [], 0
    for r in recon.itertuples():
        sh, coord = r.cell_reference.split("!")
        coord = coord.replace("$", "")
        x = xval.get((sh, coord))
        e = ev.cell(sh, coord)
        p = pym[r.item]                       # full precision, re-run in process
        rl = rel(x, p)
        ok = rl is not None and rl <= 1e-6
        if not ok:
            bad += 1
        rows.append(dict(output=r.item, cell_reference=r.cell_reference, excel_value=x,
                         python_value=p, evaluator_value=e,
                         abs_diff=(None if x is None else abs(x - p)), rel_diff=rl,
                         excel_vs_evaluator_rel=rel(x, e),
                         python_value_as_written_to_13_reconciliation_csv=float(r.python_value),
                         **{"pass": "yes" if ok else "no"}))
    pd.DataFrame(rows).to_csv(OD("17_excel_vs_python.csv"), index=False)
    print(f"[2] named outputs: {len(rows)} compared, {bad} fail at 1e-6 relative")

    # ---------------------------------------------------------------- 3. static formula review
    findings = []

    def add(sheet, cell, kind, sev, desc, fix):
        findings.append(dict(sheet=sheet, cell=cell, issue_type=kind, severity=sev,
                             description=desc, suggested_fix=fix))

    # 3a. numeric literals inside formulas
    lit = re.compile(r"(?<![A-Z0-9$.\[])(\d+(?:\.\d+)?)(?![0-9]*[A-Z(])")
    benign = {"0", "1", "2", "100", "10000", "9", "10"}
    for r in formulas.itertuples():
        stripped = REF.sub("@", r.formula)
        nums = [n for n in lit.findall(stripped) if n not in benign]
        if nums:
            add(r.sheet, r.address, "hard-coded constant in formula", "auto",
                f"literal(s) {sorted(set(nums))} in {r.formula[:150]}", "")

    # 3b. references to empty cells
    nonempty = set(xtxt)
    for r in formulas.itertuples():
        miss = []
        for m in REF.finditer(r.formula):
            sh = m.group("sq") or m.group("sheet") or r.sheet
            if m.group("c2"):
                continue                       # ranges may legitimately span blanks
            key = (sh, m.group("c1") + m.group("r1"))
            if key not in nonempty:
                miss.append(f"{sh}!{key[1]}")
        if miss:
            add(r.sheet, r.address, "reference to empty cell", "auto",
                f"{sorted(set(miss))} in {r.formula[:120]}", "")

    # 3c. inconsistent formulas along a time-series row
    fmap = {(r.sheet, r.row, r.col): r.formula for r in formulas.itertuples()}
    rowcells = {}
    for (sh, rr, cc), f in fmap.items():
        rowcells.setdefault((sh, rr), []).append((cc, f))
    for (sh, rr), cells in sorted(rowcells.items()):
        cells.sort()
        if len(cells) < 3:
            continue
        pats = {}
        for cc, f in cells:
            pats.setdefault(a1_to_r1c1(f, rr, cc), []).append(get_column_letter(cc))
        if len(pats) > 1:
            biggest = max(pats.values(), key=len)
            odd = [c for cols in pats.values() if cols is not biggest for c in cols]
            add(sh, f"row {rr}: {','.join(odd)}", "formula differs from its row neighbours", "auto",
                f"{len(pats)} distinct patterns across {[c for c, _ in [(get_column_letter(c), f) for c, f in cells]]}; "
                f"majority {biggest}", "")

    # 3d. SUM ranges that stop short of the row's populated span
    for r in formulas.itertuples():
        for m in re.finditer(r"SUM\(([^()]*)\)", r.formula):
            for rm in REF.finditer(m.group(1)):
                if not rm.group("c2"):
                    continue
                sh = rm.group("sq") or rm.group("sheet") or r.sheet
                r1, r2 = int(rm.group("r1")), int(rm.group("r2"))
                if r1 != r2:
                    continue
                c1 = column_index_from_string(rm.group("c1"))
                c2 = column_index_from_string(rm.group("c2"))
                row_pop = sorted(c for (s2, rr2, cc) in fmap if s2 == sh and rr2 == r1
                                 for c in [cc])
                if row_pop and (min(row_pop) < c1 or max(row_pop) > c2):
                    add(r.sheet, r.address, "SUM range vs populated row", "auto",
                        f"SUM {rm.group(0)} covers {get_column_letter(c1)}-{get_column_letter(c2)}; "
                        f"row {r1} of {sh} is populated "
                        f"{get_column_letter(min(row_pop))}-{get_column_letter(max(row_pop))}", "")

    pd.DataFrame(findings).to_csv(os.path.join(ROOT, "data", "processed", "overnight",
                                               "17_auto_scan.csv"), index=False)
    print(f"[3] auto scan: {len(findings)} raw hits -> data/processed/overnight/17_auto_scan.csv")
    for k, g in pd.DataFrame(findings).groupby("issue_type"):
        print(f"      {k}: {len(g)}")


if __name__ == "__main__":
    main()
