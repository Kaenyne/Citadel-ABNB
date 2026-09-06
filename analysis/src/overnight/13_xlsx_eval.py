"""A minimal Excel formula evaluator, used to verify `model/ABNB_driver_model.xlsx`.

Why this exists: LibreOffice headless is not installed on this machine and the `formulas` package is
not available, so the workbook cannot be recalculated by an external engine. This module evaluates
the subset of Excel syntax that `13_excel_builder.py` actually emits - cell and range references
(with or without $ and a sheet prefix), + - * / ^ ( ), and SUM / MIN / MAX / AVERAGE / CHOOSE - so
that `13_driver_model.py` can check every workbook output against the Python mirror for real.

It is deliberately not a general Excel engine. If the builder ever emits a function this does not
know, evaluation raises rather than guessing.

RUN (standalone spot check)   py -3.13 analysis/src/overnight/13_xlsx_eval.py
"""
from __future__ import annotations

import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

REF = re.compile(
    r"(?:(?:'(?P<sq>[^']+)'|(?P<sheet>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+))?"
)
FUNCS = {"SUM": "_sum", "MIN": "_min", "MAX": "_max", "AVERAGE": "_avg", "CHOOSE": "_choose"}
KNOWN_FUNC = re.compile(r"\b([A-Z][A-Z0-9.]*)\s*\(")


class Evaluator:
    def __init__(self, path):
        self.wb = load_workbook(path, data_only=False)
        self.cache = {}
        self.stack = []

    # -------------------------------------------------------------- helpers used inside eval()
    @staticmethod
    def _flat(args):
        out = []
        for a in args:
            out.extend(a) if isinstance(a, list) else out.append(a)
        return [x for x in out if x is not None]

    def _sum(self, *a):
        return sum(self._flat(a))

    def _min(self, *a):
        return min(self._flat(a))

    def _max(self, *a):
        return max(self._flat(a))

    def _avg(self, *a):
        f = self._flat(a)
        return sum(f) / len(f)

    def _choose(self, i, *a):
        return a[int(round(i)) - 1]

    # -------------------------------------------------------------- core
    def cell(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise RuntimeError("circular reference: " + " -> ".join(f"{s}!{c}" for s, c in self.stack + [key]))
        self.stack.append(key)
        try:
            v = self.wb[sheet][coord].value
            if isinstance(v, str) and v.startswith("="):
                out = self.evaluate(v[1:], sheet)
            elif isinstance(v, (int, float)):
                out = float(v)
            elif v is None:
                out = 0.0
            else:
                out = None                      # text label
        finally:
            self.stack.pop()
        self.cache[key] = out
        return out

    def rng(self, sheet, c1, r1, c2, r2):
        a, b = column_index_from_string(c1), column_index_from_string(c2)
        return [self.cell(sheet, f"{get_column_letter(c)}{r}")
                for c in range(min(a, b), max(a, b) + 1)
                for r in range(min(int(r1), int(r2)), max(int(r1), int(r2)) + 1)]

    def evaluate(self, formula, sheet):
        unknown = {f for f in KNOWN_FUNC.findall(formula) if f not in FUNCS}
        if unknown:
            raise NotImplementedError(f"function(s) {sorted(unknown)} not supported: {formula}")
        # protect function names from the reference regex (none of them look like refs, but be safe)
        expr = formula
        for fn, py in FUNCS.items():
            expr = re.sub(rf"\b{fn}\s*\(", f"self.{py}(", expr)

        def sub(m):
            sh = m.group("sq") or m.group("sheet") or sheet
            if m.group("c2"):
                return (f"self.rng({sh!r},{m.group('c1')!r},{m.group('r1')},"
                        f"{m.group('c2')!r},{m.group('r2')})")
            return f"self.cell({sh!r},{m.group('c1') + m.group('r1')!r})"

        # do not touch text inside self.<fn>( - the regex cannot match those, they are lowercase
        expr = REF.sub(sub, expr)
        expr = expr.replace("^", "**")
        return eval(expr, {"__builtins__": {}}, {"self": self})           # noqa: S307


def evaluate_workbook(path):
    return Evaluator(path)


if __name__ == "__main__":
    import os

    root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    ev = evaluate_workbook(os.path.join(root, "model", "ABNB_driver_model.xlsx"))
    n = bad = 0
    for ws in ev.wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    n += 1
                    try:
                        ev.cell(ws.title, c.coordinate)
                    except Exception as exc:                       # noqa: BLE001
                        bad += 1
                        print("ERROR", ws.title, c.coordinate, exc)
    print(f"{n} formula cells evaluated, {bad} errors")
