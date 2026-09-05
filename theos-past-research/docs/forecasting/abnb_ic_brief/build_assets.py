#!/usr/bin/env python3
"""Recalculate and render the auditable assets for the ABNB IC brief.

The builder reads preserved research artifacts, never modifies them, and refuses
to publish if reviewed sample sizes or reported values drift unexpectedly.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook


NAVY = "#17365D"
CHARCOAL = "#2F343B"
GREEN = "#2E7D5B"
AMBER = "#B7791F"
RED = "#B23A48"
LIGHT_GRID = "#D8DEE7"

EXPECTED_METRICS: dict[str, tuple[float | int, float]] = {
    "guidance.events": (23, 0.0),
    "guidance.numeric_ranges": (20, 0.0),
    "activity_to_guidance.n": (16, 0.0),
    "activity_to_guidance.pearson": (0.7841064501675482, 1e-10),
    "activity_to_guidance.spearman": (0.38823529411764707, 1e-10),
    "activity_to_guidance.pearson_improvement_vs_equal_weight": (
        0.00752699991030692,
        1e-10,
    ),
    "activity_to_guidance.pearson_improvement_vs_sfo": (
        0.002688896344831715,
        1e-10,
    ),
    "activity_to_guidance.acceleration_pearson": (-0.6270051347600837, 1e-10),
    "activity_to_guidance.direction_concordance": (0.5333333333333333, 1e-10),
    "activity_to_guidance.strict_pit_rows": (0, 0.0),
    "guidance_to_return.yoy_pearson": (0.07630781495666977, 1e-10),
    "guidance_to_return.yoy_n": (16, 0.0),
    "guidance_to_return.acceleration_pearson": (0.08244198563550548, 1e-10),
    "guidance_to_return.acceleration_n": (15, 0.0),
    "guidance_to_return.sequential_pct_pearson": (0.03212270033657204, 1e-10),
    "guidance_to_return.sequential_pct_n": (19, 0.0),
    "guidance_to_return.direction_aligned": (7, 0.0),
    "guidance_to_return.direction_total": (19, 0.0),
    "h10_directional.early_hits": (3, 0.0),
    "h10_directional.early_total": (4, 0.0),
    "h10_directional.later_hits": (6, 0.0),
    "h10_directional.later_total": (12, 0.0),
    "h10_directional.total_hits": (9, 0.0),
    "h10_directional.total_events": (16, 0.0),
    "controls.ose_covered_events": (4, 0.0),
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _workspace_root(reaction_workbook: Path) -> Path:
    # Expected layout: <root>/outputs/<run-id>/<workbook>.
    return reaction_workbook.resolve().parents[2]


def _reaction_frame(path: Path) -> tuple[pd.DataFrame, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    reaction_rows = list(workbook["Price Reactions"].iter_rows(values_only=True))
    header = list(reaction_rows[1])
    reactions = pd.DataFrame(reaction_rows[2:], columns=header).dropna(how="all")

    checks_rows = list(workbook["Checks"].iter_rows(values_only=True))
    checks = {
        row[0]: row[1]
        for row in checks_rows[2:]
        if row and row[0] is not None and len(row) > 1
    }
    ose_covered_events = int(checks["Historically edge-covered events"])
    workbook.close()
    return reactions, ose_covered_events


def _correlation(frame: pd.DataFrame, x: str, y: str, method: str = "pearson") -> tuple[float, int]:
    comparable = frame[[x, y]].apply(pd.to_numeric, errors="coerce").dropna()
    if len(comparable) < 2:
        return float("nan"), len(comparable)
    return float(comparable.corr(method=method).iloc[0, 1]), len(comparable)


def _h10_metrics(early_path: Path, later_path: Path) -> dict[str, int]:
    early = pd.read_csv(early_path)
    later = pd.read_csv(later_path)
    early_h10 = early.loc[early["signal_id"].eq("H-001"), "replay_classification"]
    later_h10 = later.loc[later["signal_id"].eq("H-001"), "classification"]
    early_testable = early_h10[early_h10.isin(["hit", "miss"])]
    later_testable = later_h10[later_h10.isin(["hit", "miss"])]
    return {
        "early_hits": int(early_testable.eq("hit").sum()),
        "early_total": int(len(early_testable)),
        "later_hits": int(later_testable.eq("hit").sum()),
        "later_total": int(len(later_testable)),
        "total_hits": int(early_testable.eq("hit").sum() + later_testable.eq("hit").sum()),
        "total_events": int(len(early_testable) + len(later_testable)),
    }


def _default_h10_paths(reaction_workbook: Path) -> tuple[Path, Path]:
    base = (
        _workspace_root(reaction_workbook)
        / "research"
        / "readiness"
        / "20260903T053309Z_abnb_readiness"
    )
    return (
        base / "cohort_2020q4_2023q2_replay.csv",
        base / "cohort_2023q3_2026q2_replay.csv",
    )


def compute_metrics(
    reaction_workbook: Path,
    activity_panel: Path,
    h10_early: Path | None = None,
    h10_later: Path | None = None,
) -> dict[str, Any]:
    """Compute all numeric claims displayed in the IC brief."""
    reaction_workbook = Path(reaction_workbook)
    activity_panel = Path(activity_panel)
    if h10_early is None or h10_later is None:
        h10_early, h10_later = _default_h10_paths(reaction_workbook)

    reactions, ose_covered_events = _reaction_frame(reaction_workbook)
    panel = pd.read_csv(activity_panel)
    numeric_panel_columns = [
        "feature_value",
        "feature_change_pp",
        "guidance_yoy_growth",
        "guidance_acceleration_pp",
    ]
    for column in numeric_panel_columns:
        panel[column] = pd.to_numeric(panel[column], errors="coerce")

    composite = panel.loc[
        panel["feature_id"].eq("US_EMEA_REVENUE_WEIGHTED_COMPOSITE")
    ].copy()
    equal_weight = panel.loc[panel["feature_id"].eq("US_EU_50_50_COMPOSITE")].copy()
    sfo = panel.loc[panel["feature_id"].eq("US_SFO_T3M_YOY")].copy()

    composite_pearson, composite_n = _correlation(
        composite, "feature_value", "guidance_yoy_growth"
    )
    composite_spearman, _ = _correlation(
        composite, "feature_value", "guidance_yoy_growth", "spearman"
    )
    equal_pearson, equal_n = _correlation(equal_weight, "feature_value", "guidance_yoy_growth")
    sfo_pearson, sfo_n = _correlation(sfo, "feature_value", "guidance_yoy_growth")
    if equal_n != composite_n or sfo_n != composite_n:
        raise ValueError("activity sleeves do not share a comparable sample")

    acceleration_pearson, acceleration_n = _correlation(
        composite, "feature_value", "guidance_acceleration_pp"
    )
    acceleration_spearman, _ = _correlation(
        composite, "feature_value", "guidance_acceleration_pp", "spearman"
    )
    direction = composite[["feature_change_pp", "guidance_acceleration_pp"]].dropna()
    direction_concordance = float(
        (
            direction["feature_change_pp"].ge(0)
            == direction["guidance_acceleration_pp"].ge(0)
        ).mean()
    )

    joined = composite[
        ["prediction_id", "guidance_yoy_growth", "guidance_acceleration_pp"]
    ].merge(
        reactions[["Prediction ID", "Excess Return"]],
        left_on="prediction_id",
        right_on="Prediction ID",
        how="inner",
        validate="one_to_one",
    )
    yoy_return_pearson, yoy_return_n = _correlation(
        joined, "guidance_yoy_growth", "Excess Return"
    )
    acceleration_return_pearson, acceleration_return_n = _correlation(
        joined, "guidance_acceleration_pp", "Excess Return"
    )
    sequential_pct_pearson, sequential_pct_n = _correlation(
        reactions, "\u0394 Guidance vs Prior-Q (%)", "Excess Return"
    )
    directional = reactions["Direction Comparison"].isin(["Aligned", "Diverged"])

    strict_pit_rows = int(
        composite["strict_pit_eligible"].astype(str).str.lower().eq("true").sum()
    )

    metrics: dict[str, Any] = {
        "guidance": {
            "events": int(len(reactions)),
            "numeric_ranges": int(reactions["Guidance Midpoint ($m)"].notna().sum()),
        },
        "activity_to_guidance": {
            "n": composite_n,
            "pearson": composite_pearson,
            "spearman": composite_spearman,
            "pearson_improvement_vs_equal_weight": composite_pearson - equal_pearson,
            "pearson_improvement_vs_sfo": composite_pearson - sfo_pearson,
            "acceleration_pearson": acceleration_pearson,
            "acceleration_spearman": acceleration_spearman,
            "acceleration_n": acceleration_n,
            "direction_concordance": direction_concordance,
            "direction_n": int(len(direction)),
            "strict_pit_rows": strict_pit_rows,
        },
        "guidance_to_return": {
            "yoy_pearson": yoy_return_pearson,
            "yoy_n": yoy_return_n,
            "acceleration_pearson": acceleration_return_pearson,
            "acceleration_n": acceleration_return_n,
            "sequential_pct_pearson": sequential_pct_pearson,
            "sequential_pct_n": sequential_pct_n,
            "direction_aligned": int(
                reactions.loc[directional, "Direction Comparison"].eq("Aligned").sum()
            ),
            "direction_total": int(directional.sum()),
        },
        "event_returns": {
            "n": int(reactions["Excess Return"].notna().sum()),
            "mean_abnb": float(reactions["ABNB Return"].mean()),
            "mean_excess": float(reactions["Excess Return"].mean()),
            "median_excess": float(reactions["Excess Return"].median()),
            "std_excess": float(reactions["Excess Return"].std()),
        },
        "h10_directional": _h10_metrics(Path(h10_early), Path(h10_later)),
        "controls": {"ose_covered_events": ose_covered_events},
    }
    return metrics


def _get_metric(metrics: dict[str, Any], dotted_path: str) -> float | int:
    value: Any = metrics
    for part in dotted_path.split("."):
        value = value[part]
    return value


def validate_expected_metrics(metrics: dict[str, Any]) -> None:
    """Fail closed if a report claim no longer matches the reviewed evidence."""
    mismatches = []
    for dotted_path, (expected, tolerance) in EXPECTED_METRICS.items():
        actual = _get_metric(metrics, dotted_path)
        if abs(float(actual) - float(expected)) > tolerance:
            mismatches.append(f"{dotted_path}: expected {expected!r}, found {actual!r}")
    if mismatches:
        raise ValueError("reviewed metric mismatch: " + "; ".join(mismatches))


def _tex_percent(value: float, decimals: int = 1, signed: bool = False) -> str:
    sign = "+" if signed else ""
    return f"{value * 100:{sign}.{decimals}f}\\%"


def _write_metrics_tex(metrics: dict[str, Any], path: Path) -> None:
    activity = metrics["activity_to_guidance"]
    returns = metrics["guidance_to_return"]
    event_returns = metrics["event_returns"]
    h10 = metrics["h10_directional"]
    macros = {
        "GuidanceEvents": str(metrics["guidance"]["events"]),
        "NumericGuidanceEvents": str(metrics["guidance"]["numeric_ranges"]),
        "CompositePearson": f'{activity["pearson"]:.3f}',
        "CompositeSpearman": f'{activity["spearman"]:.3f}',
        "CompositeN": str(activity["n"]),
        "CompositeLiftEqual": f'{activity["pearson_improvement_vs_equal_weight"]:+.3f}',
        "CompositeLiftSfo": f'{activity["pearson_improvement_vs_sfo"]:+.3f}',
        "CompositeAccelerationPearson": f'{activity["acceleration_pearson"]:.3f}',
        "CompositeAccelerationSpearman": f'{activity["acceleration_spearman"]:.3f}',
        "DirectionConcordance": _tex_percent(activity["direction_concordance"]),
        "StrictPitRows": str(activity["strict_pit_rows"]),
        "GuidanceReturnPearson": f'{returns["yoy_pearson"]:.3f}',
        "GuidanceReturnN": str(returns["yoy_n"]),
        "AccelerationReturnPearson": f'{returns["acceleration_pearson"]:.3f}',
        "AccelerationReturnN": str(returns["acceleration_n"]),
        "SequentialReturnPearson": f'{returns["sequential_pct_pearson"]:.3f}',
        "SequentialReturnN": str(returns["sequential_pct_n"]),
        "DirectionAlignment": (
            f'{returns["direction_aligned"]}/{returns["direction_total"]}'
        ),
        "DirectionAlignmentRate": (
            _tex_percent(returns["direction_aligned"] / returns["direction_total"])
        ),
        "MeanAbnbReturn": _tex_percent(event_returns["mean_abnb"], decimals=2, signed=True),
        "MeanExcessReturn": _tex_percent(event_returns["mean_excess"], decimals=2, signed=True),
        "MedianExcessReturn": _tex_percent(event_returns["median_excess"], decimals=2, signed=True),
        "HtenHits": f'{h10["total_hits"]}/{h10["total_events"]}',
        "HtenEarlyHits": f'{h10["early_hits"]}/{h10["early_total"]}',
        "HtenLaterHits": f'{h10["later_hits"]}/{h10["later_total"]}',
        "OseCoveredEvents": str(metrics["controls"]["ose_covered_events"]),
    }
    lines = ["% Generated by build_assets.py; do not edit by hand."]
    lines.extend(f"\\newcommand{{\\{name}}}{{{value}}}" for name, value in macros.items())
    path.write_text("\n".join(lines) + "\n", encoding="ascii")


def _style_axis(axis: plt.Axes) -> None:
    axis.spines[["top", "right"]].set_visible(False)
    axis.spines[["left", "bottom"]].set_color(LIGHT_GRID)
    axis.grid(color=LIGHT_GRID, linewidth=0.6, alpha=0.7)
    axis.tick_params(colors=CHARCOAL, labelsize=7)
    axis.set_axisbelow(True)


def _regression_line(axis: plt.Axes, x: pd.Series, y: pd.Series, color: str) -> None:
    coefficients = np.polyfit(x, y, 1)
    grid = np.linspace(float(x.min()), float(x.max()), 100)
    axis.plot(grid, coefficients[0] * grid + coefficients[1], color=color, linewidth=1.6)


def _render_correlation_contrast(
    composite: pd.DataFrame, reactions: pd.DataFrame, output_path: Path
) -> None:
    left = composite[["issuing_fiscal_period", "feature_value", "guidance_yoy_growth"]].dropna()
    right = composite[
        ["prediction_id", "issuing_fiscal_period", "guidance_yoy_growth"]
    ].merge(
        reactions[["Prediction ID", "Excess Return"]],
        left_on="prediction_id",
        right_on="Prediction ID",
        how="inner",
    ).dropna(subset=["guidance_yoy_growth", "Excess Return"])

    fig, axes = plt.subplots(1, 2, figsize=(10.2, 3.15), constrained_layout=True)
    fig.patch.set_facecolor("white")
    for axis in axes:
        _style_axis(axis)

    axes[0].scatter(
        left["feature_value"],
        left["guidance_yoy_growth"],
        color=GREEN,
        edgecolor="white",
        linewidth=0.5,
        s=28,
        zorder=3,
    )
    _regression_line(axes[0], left["feature_value"], left["guidance_yoy_growth"], GREEN)
    axes[0].set_title(
        "Activity composite -> guidance YoY\n$r$ = 0.784 | $n$ = 16",
        loc="left",
        color=NAVY,
        fontsize=9,
        fontweight="bold",
    )
    axes[0].set_xlabel("Revenue-weighted activity composite", fontsize=7.5, color=CHARCOAL)
    axes[0].set_ylabel("Guidance YoY", fontsize=7.5, color=CHARCOAL)
    axes[0].xaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axes[0].yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axes[0].text(
        0.01,
        0.98,
        "Retrospective diagnostic | zero strict-PIT rows | not alpha",
        transform=axes[0].transAxes,
        va="top",
        fontsize=6.8,
        color=RED,
        fontweight="bold",
    )

    axes[1].scatter(
        right["guidance_yoy_growth"],
        right["Excess Return"],
        color=AMBER,
        edgecolor="white",
        linewidth=0.5,
        s=28,
        zorder=3,
    )
    _regression_line(axes[1], right["guidance_yoy_growth"], right["Excess Return"], AMBER)
    axes[1].axhline(0, color=CHARCOAL, linewidth=0.8)
    axes[1].set_title(
        "Guidance YoY -> next-close excess return\n$r$ = 0.076 | $n$ = 16",
        loc="left",
        color=NAVY,
        fontsize=9,
        fontweight="bold",
    )
    axes[1].set_xlabel("Guidance YoY", fontsize=7.5, color=CHARCOAL)
    axes[1].set_ylabel("ABNB minus SPY return", fontsize=7.5, color=CHARCOAL)
    axes[1].xaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axes[1].yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    fig.savefig(output_path, format="pdf", bbox_inches="tight", metadata={"Creator": "build_assets.py"})
    plt.close(fig)


def _render_event_returns(reactions: pd.DataFrame, output_path: Path) -> None:
    plot = reactions[["Issuing Quarter", "Excess Return"]].dropna().copy()
    values = plot["Excess Return"].astype(float)
    colors = [GREEN if value >= 0 else RED for value in values]
    labels = plot["Issuing Quarter"].astype(str).str.replace("20", "", n=1)

    fig, axis = plt.subplots(figsize=(10.2, 2.45), constrained_layout=True)
    fig.patch.set_facecolor("white")
    axis.bar(np.arange(len(plot)), values, color=colors, width=0.72)
    axis.axhline(0, color=CHARCOAL, linewidth=0.8)
    axis.set_xticks(np.arange(len(plot)))
    axis.set_xticklabels(labels, rotation=55, ha="right", fontsize=6.2)
    axis.yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axis.set_ylabel("Excess return", fontsize=7.5, color=CHARCOAL)
    axis.set_title(
        "ABNB next-close excess return across 23 preserved guidance events",
        loc="left",
        fontsize=9,
        fontweight="bold",
        color=NAVY,
    )
    _style_axis(axis)
    fig.savefig(output_path, format="pdf", bbox_inches="tight", metadata={"Creator": "build_assets.py"})
    plt.close(fig)


def _render_guidance_vs_stock_performance(
    composite: pd.DataFrame, reactions: pd.DataFrame, output_path: Path
) -> None:
    """Align same-season guidance growth with benchmark-adjusted stock returns."""
    plot = composite[
        ["prediction_id", "issuing_fiscal_period", "guidance_yoy_growth"]
    ].merge(
        reactions[["Prediction ID", "Excess Return"]],
        left_on="prediction_id",
        right_on="Prediction ID",
        how="inner",
        validate="one_to_one",
    ).dropna(subset=["guidance_yoy_growth", "Excess Return"])
    labels = plot["issuing_fiscal_period"].astype(str).str.replace("20", "", n=1)
    x = np.arange(len(plot))
    excess = plot["Excess Return"].astype(float)

    fig, axes = plt.subplots(
        2,
        1,
        figsize=(10.2, 2.65),
        sharex=True,
        gridspec_kw={"height_ratios": [1.0, 1.05], "hspace": 0.08},
    )
    fig.patch.set_facecolor("white")
    axes[0].bar(x, plot["guidance_yoy_growth"], color=NAVY, width=0.68)
    axes[0].yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axes[0].set_ylabel("Guidance YoY", fontsize=7.2, color=CHARCOAL)
    axes[0].set_title(
        "Guidance growth and next-close stock performance by event",
        loc="left",
        fontsize=9,
        fontweight="bold",
        color=NAVY,
    )
    axes[0].text(
        0.995,
        0.92,
        "Pearson r = 0.076 | n = 16",
        transform=axes[0].transAxes,
        ha="right",
        va="top",
        fontsize=6.8,
        color=RED,
        fontweight="bold",
    )

    return_colors = [GREEN if value >= 0 else RED for value in excess]
    axes[1].bar(x, excess, color=return_colors, width=0.68)
    axes[1].axhline(0, color=CHARCOAL, linewidth=0.8)
    axes[1].yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axes[1].set_ylabel("ABNB - SPY", fontsize=7.2, color=CHARCOAL)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(labels, rotation=50, ha="right", fontsize=6.2)
    axes[1].set_xlabel("Issuing fiscal period", fontsize=7.2, color=CHARCOAL)
    for axis in axes:
        _style_axis(axis)

    fig.savefig(output_path, format="pdf", bbox_inches="tight", metadata={"Creator": "build_assets.py"})
    plt.close(fig)


def build_assets(
    reaction_workbook: Path,
    activity_panel: Path,
    output_dir: Path,
    h10_early: Path | None = None,
    h10_later: Path | None = None,
) -> dict[str, Any]:
    """Validate evidence, emit metrics JSON/TeX, and render vector charts."""
    reaction_workbook = Path(reaction_workbook).resolve()
    activity_panel = Path(activity_panel).resolve()
    if h10_early is None or h10_later is None:
        h10_early, h10_later = _default_h10_paths(reaction_workbook)
    h10_early = Path(h10_early).resolve()
    h10_later = Path(h10_later).resolve()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    metrics = compute_metrics(reaction_workbook, activity_panel, h10_early, h10_later)
    validate_expected_metrics(metrics)

    inputs = [reaction_workbook, activity_panel, h10_early, h10_later]
    audit = {
        "schema_version": "1.0",
        "evidence_cutoff": "2026-09-03",
        "purpose": "retrospective diagnostic and report reproducibility; not alpha",
        "inputs": [
            {"path": str(path.relative_to(_workspace_root(reaction_workbook))), "sha256": _sha256(path)}
            for path in inputs
        ],
        "metrics": metrics,
    }
    (output_dir / "metrics.json").write_text(
        json.dumps(audit, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    _write_metrics_tex(metrics, output_dir / "metrics.tex")

    reactions, _ = _reaction_frame(reaction_workbook)
    panel = pd.read_csv(activity_panel)
    for column in ("feature_value", "guidance_yoy_growth"):
        panel[column] = pd.to_numeric(panel[column], errors="coerce")
    composite = panel.loc[
        panel["feature_id"].eq("US_EMEA_REVENUE_WEIGHTED_COMPOSITE")
    ].copy()
    _render_correlation_contrast(composite, reactions, output_dir / "correlation_contrast.pdf")
    _render_guidance_vs_stock_performance(
        composite, reactions, output_dir / "guidance_vs_stock_performance.pdf"
    )
    _render_event_returns(reactions, output_dir / "event_excess_returns.pdf")
    return audit


def main() -> int:
    workspace = Path(__file__).resolve().parents[3]
    # Task 5 consolidates reviewable inputs in stable, named output locations.
    default_reaction_workbook = (
        workspace / "outputs" / "workbooks" / "ABNB_edge_guidance_stock_reaction.xlsx"
    )
    default_activity_panel = (
        workspace
        / "outputs"
        / "reproducibility"
        / "us-europe-guidance"
        / "abnb_us_europe_guidance_panel.csv"
    )
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--reaction-workbook",
        type=Path,
        default=default_reaction_workbook,
    )
    parser.add_argument(
        "--activity-panel",
        type=Path,
        default=default_activity_panel,
    )
    parser.add_argument(
        "--output-dir", type=Path, default=Path(__file__).resolve().parent / "generated"
    )
    args = parser.parse_args()
    build_assets(args.reaction_workbook, args.activity_panel, args.output_dir)
    print(f"ABNB IC brief assets written to {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
